# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- palette ----
NAVY="1F2A44"; BLUE="2E5AAC"; TEAL="0E7C7B"; ORANGE="C05621"; PURPLE="5B3A8C"
LT_BLUE="E8EEF9"; LT_GREY="F4F5F7"; MID_GREY="DDE1E6"; WHITE="FFFFFF"; YELLOW="FFF3C4"
GREEN="1E7D34"; RED="B02A37"
FONT="Calibri"

thin=Side(style="thin", color="C9CED6")
border=Border(left=thin,right=thin,top=thin,bottom=thin)

def style_header(ws, row, ncols, color=NAVY):
    for c in range(1,ncols+1):
        cell=ws.cell(row=row,column=c)
        cell.font=Font(name=FONT,bold=True,color="FFFFFF",size=11)
        cell.fill=PatternFill("solid",fgColor=color)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=border

def title_block(ws, text, sub, ncols):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
    t=ws.cell(row=1,column=1,value=text)
    t.font=Font(name=FONT,bold=True,size=16,color="FFFFFF")
    t.fill=PatternFill("solid",fgColor=NAVY)
    t.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[1].height=30
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncols)
    s=ws.cell(row=2,column=1,value=sub)
    s.font=Font(name=FONT,italic=True,size=10,color="333333")
    s.fill=PatternFill("solid",fgColor=LT_GREY)
    s.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[2].height=20

RUPEE='"₹"#,##0'

wb=Workbook()

# ============================================================= OVERVIEW
ws=wb.active; ws.title="Overview"
ws.sheet_view.showGridLines=False
title_block(ws,"HARSHDEEP GROUP — 90-Day Scale Engine","Aurra Hype · Designomics India · Loop In Events  |  Powered in-house by Arihant Digital  |  Prepared Jul 2026",2)
rows=[
 ("THE THESIS",""),
 ("One founder, four assets, one flywheel","You already own a ~10-yr digital agency (Arihant Digital). Treat it as the in-house engine that builds & markets the three new brands at cost — then sell that same capability to outside clients as proof accumulates."),
 ("Aurra Hype","D2C streetwear / hype apparel (Bengaluru). Physical store + office launched 12 Jul. Sells online via aurahyped.com. Growth = product drops + performance ads + marketplace + quick-commerce."),
 ("Designomics India","Personalised gifting + merch + apparel (custom boxes, stickers, mugs, oversized tees, workshops, corporate gifting). Growth = corporate gifting contracts + custom-box e-comm + B2B2C."),
 ("Loop In Events","Event management & experiential. NAME CONFLICT: 'Loop In Events' + loopinevents.com already belong to an established Canadian planner (Michelle Munro). Rebrand + trademark strongly advised — see Roadmap & Doc."),
 ("Arihant Digital (existing)","The unfair advantage. Supplies web, performance marketing, SEO/GEO, automation & analytics to all three at internal cost, then externalises as a productised agency offer."),
 ("",""),
 ("HOW TO READ THIS WORKBOOK",""),
 ("Master Stack","Every service, software & automation the group needs, tagged by business, priority, in-house-vs-buy, vendor and cost."),
 ("AI Automations","Specific automations to build (trigger → action → tools → impact)."),
 ("90-Day Roadmap","Week-by-week plan across all three brands with the 'proof' KPI for each phase."),
 ("Budget Summary","Setup + monthly spend rolled up by domain (aggressive-spend scenario)."),
 ("Q-Commerce & Marketplaces","Channel-by-channel onboarding tracker with real 2026 India costs & fit."),
 ("",""),
 ("LEGEND",""),
 ("Priority","P0 = do in Month 1 (foundation)  ·  P1 = Month 2 (engine)  ·  P2 = Month 3 (scale)"),
 ("In-house vs Buy","In-house: Arihant = digital/marketing  ·  In-house: Designomics = design/print/merch  ·  Buy = external SaaS/vendor  ·  Hybrid = both"),
 ("Costs","INR. 'Aggressive spend to scale' scenario. Ad budgets are separate line items, not tool costs."),
]
r=4
for a,b in rows:
    ca=ws.cell(row=r,column=1,value=a); cb=ws.cell(row=r,column=2,value=b)
    is_head = (b=="" and a==a.upper() and a!="")
    if is_head:
        ca.font=Font(name=FONT,bold=True,size=12,color=NAVY)
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
        ca.fill=PatternFill("solid",fgColor=LT_BLUE)
    else:
        ca.font=Font(name=FONT,bold=True,size=10,color=BLUE)
        cb.font=Font(name=FONT,size=10,color="222222")
    ca.alignment=Alignment(vertical="top",wrap_text=True,horizontal="left")
    cb.alignment=Alignment(vertical="top",wrap_text=True,horizontal="left")
    r+=1
ws.column_dimensions["A"].width=26
ws.column_dimensions["B"].width=104

# ============================================================= MASTER STACK
ms=wb.create_sheet("Master Stack")
ms.sheet_view.showGridLines=False
headers=["#","Business","Domain","Item / Capability","What it does & why it matters","Type","Priority","In-house vs Buy","Recommended Tool / Vendor","Setup (₹)","Monthly (₹)","Phase"]
# each row: Business, Domain, Item, What, Type, Priority, InHouse, Vendor, Setup, Monthly, Phase
S=[
# --- BRAND & IDENTITY ---
("Group","Brand & Identity","Group holding brand + naming architecture","Define parent brand + how 3 sub-brands relate; fixes messy, inconsistent identity","Service","P0","In-house: Designomics","Designomics design team","40000",0,"M1"),
("Loop In","Brand & Identity","Rebrand (new name + logo + identity)","Resolve Canada name clash; secure a distinct, trademarkable name","Service","P0","In-house: Designomics","Designomics + IP attorney","60000",0,"M1"),
("Aurra Hype","Brand & Identity","Brand guideline + logo refinement","Distinguish from many 'Aura/Aurra' apparel brands; consistent look","Service","P0","In-house: Designomics","Designomics design team","30000",0,"M1"),
("Designomics","Brand & Identity","Brand system for gifting vs apparel lines","Separate but linked identity for gifting box vs streetwear tees","Service","P1","In-house: Designomics","Designomics design team","20000",0,"M2"),
("Group","Brand & Identity","Trademark search + filing (per brand, key classes)","Legal ownership of names/logos; blocks copycats","Service","P0","Buy","IP attorney / Vakilsearch / IPIndia","45000",0,"M1"),
# --- WEBSITE & ECOMMERCE ---
("Aurra Hype","Website & E-commerce","Shopify store (migrate/upgrade)","Fast, conversion-optimised storefront + apps ecosystem","SaaS","P0","In-house: Arihant","Shopify (Basic→Growth)","25000","3500","M1"),
("Designomics","Website & E-commerce","Store upgrade / replatform from Wix","Wix limits apps, speed, custom-box UX; move to Shopify","SaaS","P1","In-house: Arihant","Shopify + custom-box app","30000","3500","M2"),
("Loop In","Website & E-commerce","Website + lead-capture + portfolio","Booking enquiries, case studies, packages","Service","P0","In-house: Arihant","Webflow / WordPress","25000","1500","M1"),
("Group","Website & E-commerce","Custom domains, SSL, email, DNS","Professional infra across brands","Infra","P0","In-house: Arihant","Google Workspace + Cloudflare","5000","3000","M1"),
("Aurra Hype","Website & E-commerce","Conversion apps (reviews, upsell, wishlist)","Higher AOV & trust: Judge.me, ReConvert, Loox","SaaS","P1","Buy","Judge.me / Loox / ReConvert",0,"4000","M2"),
("Aurra Hype","Website & E-commerce","Site speed + CRO audit","Cut bounce, lift conversion on paid traffic","Service","P1","In-house: Arihant","Arihant CRO","15000",0,"M2"),
# --- CONTENT & CREATIVE ---
("Group","Content & Creative","Product & lifestyle photography","Sales-grade catalog + ad creative; current visuals weak","Service","P0","Hybrid","In-house shoot + freelance photographer","60000","20000","M1"),
("Group","Content & Creative","Short-form video / reels engine","3-5 reels/week per brand for organic + ads","Service","P0","In-house: Arihant","In-house content pod + CapCut Pro","20000","15000","M1"),
("Group","Content & Creative","Design subscription (creatives, ads, packaging art)","Unlimited on-brand creatives at speed","In-house: Designomics","P0","In-house: Designomics","Designomics + Canva Teams / Adobe CC","0","6000","M1"),
("Group","Content & Creative","UGC creator pipeline","Authentic content that converts on Meta/IG","Service","P1","Buy","UGC creators / Trend / local","0","40000","M2"),
("Aurra Hype","Content & Creative","Drop 'lookbook' + campaign shoots","Hype cycle around each product drop","Service","P1","Hybrid","Designomics + models","40000","25000","M2"),
# --- PERFORMANCE MARKETING ---
("Aurra Hype","Performance Marketing","Meta Ads (IG/FB) — always-on","Primary revenue driver for D2C apparel","Service+Spend","P0","In-house: Arihant","Meta Ads Manager","0","150000","M1"),
("Aurra Hype","Performance Marketing","Google Ads (Search+Shopping+PMax)","Capture high-intent buyers + Shopping feed","Service+Spend","P1","In-house: Arihant","Google Ads + Merchant Center","0","80000","M2"),
("Designomics","Performance Marketing","Meta + Google (gifting + corporate intent)","Seasonal gifting demand + corporate keywords","Service+Spend","P1","In-house: Arihant","Meta + Google","0","80000","M2"),
("Loop In","Performance Marketing","Local + LinkedIn lead-gen","B2B event enquiries + local weddings/corporate","Service+Spend","P1","In-house: Arihant","Google Local + LinkedIn Ads","0","40000","M2"),
("Group","Performance Marketing","Creative testing framework","Systematic winner-finding; lower CAC","Service","P1","In-house: Arihant","Arihant + Motion/Triple Whale","0","8000","M2"),
# --- ORGANIC & SOCIAL ---
("Group","Organic & Social","Social media management + calendar","Consistent posting; current presence is thin/irregular","Service","P0","In-house: Arihant","Metricool / Buffer + Arihant","0","4000","M1"),
("Group","Organic & Social","Community + DM management","Fast replies convert followers to buyers","Service","P0","In-house: Arihant","In-house + WhatsApp/IG inbox","0","6000","M1"),
("Aurra Hype","Organic & Social","Influencer seeding (micro + mid)","Hype & reach for streetwear audience","Service+Spend","P1","Buy","Influencer.in / local agencies","0","100000","M2"),
# --- SEO / GEO / AEO ---
("Group","SEO / GEO / AEO","Technical + on-page SEO","Long-term organic traffic across brands","Service","P1","In-house: Arihant","Ahrefs/SEMrush + Arihant","20000","9000","M2"),
("Group","SEO / GEO / AEO","GEO / AEO (AI-answer optimisation)","Get cited by ChatGPT/Gemini/Perplexity/AI Overviews","Service","P1","In-house: Arihant","Schema + Arihant GEO playbook","15000","5000","M2"),
("Group","SEO / GEO / AEO","Google Business Profiles + local SEO","Store footfall (Aurra store) + local event leads","Service","P0","In-house: Arihant","Google Business Profile","0","2000","M1"),
("Group","SEO / GEO / AEO","Content / blog engine","Ranking + AI-citation surface; topical authority","Service","P2","In-house: Arihant","Arihant + Surfer/Frase","0","6000","M3"),
# --- CRM & RETENTION ---
("Aurra Hype","CRM & Retention","Email + SMS + WhatsApp flows","Abandoned cart, win-back, post-purchase = repeat revenue","SaaS","P0","In-house: Arihant","Klaviyo / WebEngage","0","9000","M1"),
("Group","CRM & Retention","WhatsApp Business API + broadcast","#1 India channel: order updates, drops, offers","SaaS","P0","Buy","Interakt / WATI / AiSensy","5000","5000","M1"),
("Group","CRM & Retention","Loyalty + referral program","Repeat purchase + word-of-mouth CAC cut","SaaS","P1","Buy","Zithara / Loox referrals","0","4000","M2"),
("Loop In","CRM & Retention","Sales CRM (pipeline for enquiries)","Track event leads → quote → booking","SaaS","P1","Buy","Zoho CRM / HubSpot Free","0","2500","M2"),
# --- AI & AUTOMATION ---
("Group","AI & Automation","Automation backbone (workflows)","Connect store, ads, CRM, sheets, WhatsApp","SaaS","P0","In-house: Arihant","n8n (self-host) / Make","6000","3000","M1"),
("Group","AI & Automation","AI customer-support chatbot","24x7 FAQ, order status, sizing on site + WhatsApp","SaaS","P1","In-house: Arihant","Chatbase / Intercom Fin","0","6000","M2"),
("Group","AI & Automation","AI content & creative assist","Copy, captions, product descriptions, variations","SaaS","P0","In-house: Arihant","ChatGPT Team / Claude","0","5000","M1"),
("Group","AI & Automation","AI product photography / model shots","On-model images without full shoots","SaaS","P1","In-house: Designomics","Flair / Photoroom / Midjourney","0","4000","M2"),
("Group","AI & Automation","Ops copilots & dashboards","Auto daily sales/ads/inventory digest to WhatsApp","Automation","P1","In-house: Arihant","n8n + Sheets + GPT","0","1000","M2"),
# --- CUSTOMER SUPPORT ---
("Group","Customer Support","Helpdesk / shared inbox","One place for email, IG, WhatsApp, chat tickets","SaaS","P1","Buy","Gorgias / Zoho Desk","0","6000","M2"),
("Group","Customer Support","Returns / exchange management","Streamlined apparel returns; protect margin","SaaS","P1","Buy","Shiprocket / Return Prime","0","3000","M2"),
# --- SOURCING & MANUFACTURING ---
("Aurra Hype","Sourcing & Manufacturing","Apparel manufacturing partners","Reliable blanks + custom cut-and-sew for tees/hoodies","Service","P0","Buy","Tiruppur / Ludhiana / Bengaluru units","0","0","M1"),
("Aurra Hype","Sourcing & Manufacturing","Print / DTG / DTF partner","Small-batch prints for drops without big MOQs","Service","P0","Hybrid","Designomics + Qikink / Blinkstore","0","0","M1"),
("Designomics","Sourcing & Manufacturing","Merch & gifting supplier network","Mugs, bottles, candles, notebooks, kits at scale","Service","P0","Buy","Sourcing agents / IndiaMART verified","0","0","M1"),
("Group","Sourcing & Manufacturing","Print-on-demand fallback","Zero-inventory long-tail SKUs","Service","P1","Buy","Qikink / Blinkstore / Printrove","0","0","M2"),
("Aurra Hype","Sourcing & Manufacturing","Quality control + tech-pack system","Consistent sizing/quality; fewer returns","Service","P1","In-house: Designomics","Tech packs + QC checklist","10000",0,"M2"),
# --- PACKAGING & UNBOXING ---
("Group","Packaging & Unboxing","Custom branded packaging (boxes, mailers, tape)","Unboxing = free marketing + perceived value","Service","P0","In-house: Designomics","Designomics design + Bizongo/Pack','","35000","25000","M1"),
("Designomics","Packaging & Unboxing","Custom gift-box engine","Core product: build-your-own box UX + inserts","Service","P0","In-house: Designomics","Designomics + box configurator","20000",0,"M1"),
("Group","Packaging & Unboxing","Inserts, thank-you cards, stickers, QR","Reviews, referrals, repeat-order QR codes","Service","P1","In-house: Designomics","Designomics print","0","8000","M2"),
# --- INVENTORY & WAREHOUSE ---
("Group","Inventory & Warehouse","Inventory management system","Single source of stock across web + marketplaces + store","SaaS","P0","Buy","Zoho Inventory / Unicommerce","0","5000","M1"),
("Group","Inventory & Warehouse","Barcode / SKU + label system","Scan-based picking; stops overselling","SaaS","P1","Buy","Zoho / QuickBooks + barcode scanner","8000","1500","M2"),
("Group","Inventory & Warehouse","Warehouse / stockroom setup + WMS-lite","Organised bins for store+online fulfilment","Service","P1","Buy","Racking + Zoho Inventory","50000","0","M2"),
("Group","Inventory & Warehouse","Multi-channel stock sync","Auto-sync stock to Amazon/Flipkart/Q-comm","SaaS","P1","Buy","Unicommerce / Vinculum","0","8000","M2"),
# --- ORDER & LOGISTICS ---
("Group","Order & Logistics","Shipping aggregator","Cheapest courier per pincode; NDR management","SaaS","P0","Buy","Shiprocket / iThink Logistics","0","4000","M1"),
("Group","Order & Logistics","COD + prepaid + RTO control","Reduce RTO losses (big apparel killer)","SaaS","P1","Buy","Shiprocket COD suite / GoKwik","0","5000","M2"),
("Group","Order & Logistics","Order management system (OMS)","Unified orders from all channels in one queue","SaaS","P1","Buy","Unicommerce / Zoho","0","6000","M2"),
("Group","Order & Logistics","Fast checkout (1-click + UPI)","Cut checkout drop-off; UPI/COD in one tap","SaaS","P1","Buy","GoKwik / Shopflo / Razorpay Magic","0","5000","M2"),
("Group","Order & Logistics","Hyperlocal same-day (Bengaluru)","Same-day delivery for local orders + store","Service","P2","Buy","Dunzo/Porter/Shiprocket Quick","0","4000","M3"),
# --- QUICK COMMERCE & MARKETPLACES ---
("Aurra Hype","Q-Commerce & Marketplaces","Amazon + Flipkart apparel listings","Massive discovery; A+ content + Cataloguing","Service","P1","In-house: Arihant","Seller Central + Flipkart","0","10000","M2"),
("Aurra Hype","Q-Commerce & Marketplaces","Myntra / Ajio (fashion marketplaces)","Category-native fashion buyers","Service","P2","Buy","Myntra M-Now / Ajio onboarding","0","0","M3"),
("Designomics","Q-Commerce & Marketplaces","Quick-commerce (gifting SKUs)","Impulse gifting via Blinkit/Zepto/Instamart","Service+Fee","P2","Buy","Blinkit/Zepto/Instamart","100000","250000","M3"),
("Group","Q-Commerce & Marketplaces","Marketplace ads (Amazon/Flipkart)","Sponsored products to win Buy Box/visibility","Service+Spend","P2","In-house: Arihant","Amazon Ads / Flipkart Ads","0","40000","M3"),
# --- RETAIL / OFFLINE ---
("Aurra Hype","Retail / Offline","Retail POS + store ops","Unify store + online inventory & billing","SaaS","P0","Buy","Shopify POS / GoFrugal","10000","3000","M1"),
("Aurra Hype","Retail / Offline","Store launch event (12 Jul) amplification","Turn launch into content + PR + footfall","Service","P0","In-house: Arihant","Arihant + Loop In events team","50000",0,"M1"),
("Group","Retail / Offline","Pop-ups + exhibitions + college activations","Direct sales + UGC + email capture","Service","P2","Hybrid","Loop In + Designomics","40000","30000","M3"),
# --- FINANCE & COMPLIANCE ---
("Group","Finance & Compliance","Accounting + GST software","Clean books, GST filing across entities","SaaS","P0","Buy","Zoho Books / TallyPrime","6000","2500","M1"),
("Group","Finance & Compliance","Payment gateway + reconciliation","Accept UPI/cards/COD; auto-reconcile payouts","SaaS","P0","Buy","Razorpay / Cashfree","0","0","M1"),
("Group","Finance & Compliance","CA / bookkeeping + GST filing","Compliance for 3 brands; investor-ready books","Service","P0","Buy","CA firm (retainer)","0","15000","M1"),
("Group","Finance & Compliance","Business structure + entity setup","LLP/Pvt Ltd, GST, trademarks, IEC for exports","Service","P0","Buy","CA + CS","40000",0,"M1"),
("Group","Finance & Compliance","Unit economics + P&L dashboard","Know CAC, contribution margin, RTO per brand","Service","P1","In-house: Arihant","Sheets/Looker + Arihant","0","2000","M2"),
# --- LEGAL & IP ---
("Group","Legal & IP","Trademark portfolio management","Own names/logos in relevant classes; renewals","Service","P0","Buy","IP attorney retainer","0","6000","M1"),
("Group","Legal & IP","Contracts, policies, vendor agreements","Supplier, influencer, employment, T&Cs","Service","P1","Buy","Legal retainer / Vakilsearch","0","6000","M2"),
# --- HR & TEAM OPS ---
("Group","HR & Team Ops","Core team hiring (marketing, ops, design)","People to run the machine as it scales","Service","P1","Hybrid","Naukri/LinkedIn + referrals","0","0","M2"),
("Group","HR & Team Ops","Payroll + attendance","Compliant salaries across brands","SaaS","P1","Buy","RazorpayX Payroll / Keka","0","3000","M2"),
("Group","HR & Team Ops","SOPs + training library","Repeatable ops; onboard fast","Service","P1","In-house: Arihant","Notion + Loom","0","1500","M2"),
# --- PROJECT & INTERNAL OPS ---
("Group","Project & Internal Ops","Project / task management","One system for all 3 brands' workstreams","SaaS","P0","In-house: Arihant","Notion / ClickUp","0","3000","M1"),
("Group","Project & Internal Ops","Central docs + asset library (DAM)","Brand assets, contracts, SOPs in one place","SaaS","P1","In-house: Arihant","Google Drive / Notion","0","1500","M2"),
("Loop In","Project & Internal Ops","Event project + vendor management","Timelines, vendors, checklists per event","SaaS","P1","Buy","Monday.com / Notion + Aisle Planner","0","4000","M2"),
# --- ANALYTICS & BI ---
("Group","Analytics & BI","Web + funnel analytics","GA4, server-side tracking, attribution","SaaS","P0","In-house: Arihant","GA4 + Meta CAPI + Stape","5000","2000","M1"),
("Aurra Hype","Analytics & BI","E-com profit analytics","True profit after ads/RTO/COGS per SKU","SaaS","P1","Buy","TripleWhale / Lifetimely","0","7000","M2"),
("Group","Analytics & BI","Group KPI dashboard","One board: revenue, CAC, AOV, ROAS per brand","Service","P1","In-house: Arihant","Looker Studio + Arihant","0","2000","M2"),
]
# write master stack
ms.merge_cells("A1:L1")
c=ms.cell(row=1,column=1,value="MASTER STACK — Services · Software · Automations (all 3 brands)")
c.font=Font(name=FONT,bold=True,size=15,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=NAVY)
c.alignment=Alignment(horizontal="left",vertical="center",indent=1); ms.row_dimensions[1].height=28
ms.merge_cells("A2:L2")
c=ms.cell(row=2,column=1,value="Aggressive-spend scenario · Ad budgets shown as monthly line items · In-house = built by Arihant/Designomics at internal cost")
c.font=Font(name=FONT,italic=True,size=9,color="333333"); c.fill=PatternFill("solid",fgColor=LT_GREY)
c.alignment=Alignment(horizontal="left",indent=1); ms.row_dimensions[2].height=16
hr=4
for i,h in enumerate(headers,1): ms.cell(row=hr,column=i,value=h)
style_header(ms,hr,len(headers))
biz_color={"Aurra Hype":"FCE4D6","Designomics":"E2EFDA","Loop In":"FFF2CC","Group":"E8EEF9"}
pri_color={"P0":RED,"P1":ORANGE,"P2":TEAL}
r=hr+1
for i,row in enumerate(S,1):
    biz,dom,item,what,typ,pri,inh,ven,setup,mon,ph=row
    vals=[i,biz,dom,item,what,typ,pri,inh,ven,int(setup),(int(mon) if str(mon).isdigit() else mon),ph]
    for ci,v in enumerate(vals,1):
        cell=ms.cell(row=r,column=ci,value=v)
        cell.border=border
        cell.font=Font(name=FONT,size=9,color="222222")
        cell.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if ci in(1,6,7,10,11,12) else "left"))
    ms.cell(row=r,column=2).fill=PatternFill("solid",fgColor=biz_color.get(biz,"FFFFFF"))
    pc=ms.cell(row=r,column=7); pc.font=Font(name=FONT,size=9,bold=True,color="FFFFFF")
    pc.fill=PatternFill("solid",fgColor=pri_color.get(pri,"999999"))
    ms.cell(row=r,column=10).number_format=RUPEE
    ms.cell(row=r,column=11).number_format=RUPEE
    if r%2==0:
        for ci in [1,3,4,5,6,8,9,12]:
            if ms.cell(row=r,column=ci).fill.fgColor.rgb in (None,"00000000"):
                ms.cell(row=r,column=ci).fill=PatternFill("solid",fgColor="FAFBFC")
    r+=1
# totals row
tot=r
ms.cell(row=tot,column=9,value="TOTALS →").font=Font(name=FONT,bold=True,size=10,color=NAVY)
ms.cell(row=tot,column=9).alignment=Alignment(horizontal="right")
ms.cell(row=tot,column=10,value=f"=SUM(J{hr+1}:J{tot-1})").number_format=RUPEE
ms.cell(row=tot,column=11,value=f"=SUM(K{hr+1}:K{tot-1})").number_format=RUPEE
for ci in (10,11):
    ms.cell(row=tot,column=ci).font=Font(name=FONT,bold=True,size=10,color=NAVY)
    ms.cell(row=tot,column=ci).fill=PatternFill("solid",fgColor=YELLOW)
ms.freeze_panes="A5"
widths=[4,12,20,30,46,11,9,20,26,12,13,7]
for i,w in enumerate(widths,1): ms.column_dimensions[get_column_letter(i)].width=w
ms.auto_filter.ref=f"A{hr}:L{tot-1}"

wb.save("/sessions/bold-epic-ramanujan/mnt/outputs/Harshdeep_Group_Stack.xlsx")
print("part1 saved", len(S), "stack rows")
