# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F2A44"; BLUE="2E5AAC"; TEAL="0E7C7B"; ORANGE="C05621"; PURPLE="5B3A8C"; PLUM="7A2E5B"
LT_BLUE="E8EEF9"; LT_GREY="F4F5F7"; YELLOW="FFF3C4"; GREEN="1E7D34"; RED="B02A37"
FONT="Calibri"; RUPEE='"₹"#,##0'; USD='"$"#,##0'
thin=Side(style="thin",color="C9CED6"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
BIZC={"Aurra Hype":"FCE4D6","Designomics":"E2EFDA","Loop In":"FFF2CC","Arihant":"E1D5F0","Group":"E8EEF9","All":"E8EEF9"}
PRI={"P0":RED,"P1":ORANGE,"P2":TEAL,"P3":"7A8288"}

wb=Workbook()

def sheet(name):
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False; return ws
def banner(ws,title,sub,ncols,color=NAVY):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
    c=ws.cell(row=1,column=1,value=title); c.font=Font(name=FONT,bold=True,size=15,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor=color); c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[1].height=28
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncols)
    c=ws.cell(row=2,column=1,value=sub); c.font=Font(name=FONT,italic=True,size=9,color="333333")
    c.fill=PatternFill("solid",fgColor=LT_GREY); c.alignment=Alignment(horizontal="left",indent=1); ws.row_dimensions[2].height=16
def hdr(ws,row,headers,color=NAVY):
    for i,h in enumerate(headers,1): ws.cell(row=row,column=i,value=h)
    for cc in range(1,len(headers)+1):
        cell=ws.cell(row=row,column=cc); cell.font=Font(name=FONT,bold=True,color="FFFFFF",size=9)
        cell.fill=PatternFill("solid",fgColor=color); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
def writerows(ws,start,rows,aligns,alt=True,fontsize=9):
    r=start
    for row in rows:
        for ci,v in enumerate(row,1):
            cell=ws.cell(row=r,column=ci,value=v); cell.border=border
            cell.font=Font(name=FONT,size=fontsize,color="222222")
            cell.alignment=Alignment(vertical="top",wrap_text=True,horizontal=aligns[ci-1])
        if alt and r%2==0:
            for ci in range(1,len(row)+1): ws.cell(row=r,column=ci).fill=PatternFill("solid",fgColor="FAFBFC")
        r+=1
    return r
def widths(ws,ws_widths):
    for i,w in enumerate(ws_widths,1): ws.column_dimensions[get_column_letter(i)].width=w

# ================================================================ OVERVIEW
ws=wb.active; ws.title="Overview"; ws.sheet_view.showGridLines=False
banner(ws,"HARSHDEEP GROUP — MASTER OPERATING STACK (FULL / ADVANCED)","4 brands · world-class + advanced global tooling · AI-native · Founder OS · Prepared Jul 2026",2)
rows=[
 ("THE 4 ASSETS",""),
 ("Aurra Hype","D2C streetwear / hype apparel (Bengaluru). Store + office launch 12 Jul. aurahyped.com."),
 ("Designomics India","Personalised gifting + merch + apparel (custom boxes, stickers, mugs, tees, workshops, corporate gifting). Also the group's in-house design/print/packaging arm."),
 ("Loop In Events","Event management & experiential. NAME/DOMAIN CLASH with a Canadian firm (loopinevents.com) — rebrand + trademark first. Also the group's live-activation arm."),
 ("Arihant Digital","~10-yr digital agency = the in-house engine (web, performance, SEO/GEO, automation, analytics). Builds the 3 brands at cost, then sells the proven capability to outside clients."),
 ("",""),
 ("WHAT'S NEW IN THIS FULL VERSION",""),
 ("Advanced global stack","Best-in-class tools world-class brands actually use (Shopify, Klaviyo, Triple Whale, Motion, Attentive, Gorgias, GRIN, etc.) added across every function."),
 ("AI & creative engine","Full 2026 AI stack: Higgsfield (UGC ad video), Veo 3.1 / Runway / Kling (video), Midjourney / Krea / Ideogram (image), ElevenLabs / HeyGen (voice+avatar), Claude + MCP, Gemini."),
 ("Chat & social automation","IG comment-to-DM, WhatsApp API, Telegram, Facebook/Messenger via ManyChat + AiSensy + n8n agents."),
 ("Founder OS","A single command center to run all 4 brands: dashboards, daily/weekly/monthly cadence, founder-level automations & AI copilots. See 'Founder OS' tab + separate playbook doc."),
 ("Per-brand tabs","Dedicated deep tab for each of the 4 brands with its own priority stack."),
 ("World-class benchmarks","'Benchmarks' tab: what leading global brands use, by function, so you copy proven playbooks."),
 ("",""),
 ("HOW TO READ",""),
 ("Priority","P0 = Month 1 (foundation) · P1 = Month 2 (engine) · P2 = Month 3 (scale) · P3 = later / advanced."),
 ("In-house vs Buy","In-house: Arihant (digital/marketing) · In-house: Designomics (design/print/merch) · Buy (SaaS/vendor) · Hybrid."),
 ("Costs","INR unless a tool is billed in USD (AI tools tab shows USD). Aggressive-spend scenario. Ad budgets are separate line items."),
]
r=4
for a,b in rows:
    ca=ws.cell(row=r,column=1,value=a); cb=ws.cell(row=r,column=2,value=b)
    if b=="" and a:
        ca.font=Font(name=FONT,bold=True,size=12,color=NAVY); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
        ca.fill=PatternFill("solid",fgColor=LT_BLUE)
    else:
        ca.font=Font(name=FONT,bold=True,size=10,color=BLUE); cb.font=Font(name=FONT,size=10,color="222222")
    ca.alignment=Alignment(vertical="top",wrap_text=True); cb.alignment=Alignment(vertical="top",wrap_text=True)
    r+=1
widths(ws,[26,104])

# ================================================================ MASTER STACK
ms=sheet("Master Stack")
banner(ms,"MASTER STACK — every service · software · automation (advanced, all 4 brands)","P0→P3 priority · in-house vs buy · vendor · cost · phase. Filterable.",12)
H=["#","Business","Domain","Item / Capability","What it does & why it matters","Type","Priority","In-house vs Buy","Recommended Tool / Vendor","Setup (₹)","Monthly (₹)","Phase"]
hdr(ms,4,H)
S=[
# Brand & Identity
("Group","Brand & Identity","Group holding-brand & naming architecture","How the 4 sub-brands relate; fixes inconsistent identity","Service","P0","In-house: Designomics","Designomics","40000",0,"M1"),
("Loop In","Brand & Identity","Rebrand (name+logo+identity)","Resolve Canada clash; trademarkable name","Service","P0","In-house: Designomics","Designomics + IP attorney","60000",0,"M1"),
("Aurra Hype","Brand & Identity","Identity refinement + trademark","Stand out from crowded 'Aura/Aurra' space","Service","P0","In-house: Designomics","Designomics","30000",0,"M1"),
("Group","Brand & Identity","Trademark search + filing (key classes)","Legal ownership; blocks copycats","Service","P0","Buy","IP attorney / IPIndia","45000",0,"M1"),
("Group","Brand & Identity","Brand voice + messaging system","Consistent tone across all channels & AI prompts","Service","P1","In-house: Arihant","Brand book + GPT/Claude style guide","15000",0,"M2"),
# Website & Ecommerce
("Aurra Hype","Website & E-commerce","Shopify store (headless-ready)","Fast conversion storefront + app ecosystem","SaaS","P0","In-house: Arihant","Shopify → Plus later","25000","3500","M1"),
("Designomics","Website & E-commerce","Replatform off Wix + custom-box builder","Configurator UX, apps, speed","SaaS","P1","In-house: Arihant","Shopify + Kickflip/Zakeke","30000","5000","M2"),
("Loop In","Website & E-commerce","Lead-gen site + portfolio + booking","Enquiries, case studies, packages","Service","P0","In-house: Arihant","Webflow / WordPress","25000","1500","M1"),
("Group","Website & E-commerce","Landing-page / funnel builder","Fast campaign & drop landing pages","SaaS","P1","In-house: Arihant","Replo / Unbounce","0","4000","M2"),
("Aurra Hype","Website & E-commerce","On-site personalization + upsell","Higher AOV via smart recommendations","SaaS","P2","Buy","Rebuy / Nosto","0","8000","M3"),
("Group","Website & E-commerce","A/B testing & CRO","Systematic conversion lift on paid traffic","SaaS","P2","In-house: Arihant","VWO / Intelligems","0","9000","M3"),
("Group","Website & E-commerce","Domains, SSL, email, DNS, CDN","Professional infra across brands","Infra","P0","In-house: Arihant","Google Workspace + Cloudflare","5000","3000","M1"),
# Content & Creative
("Group","Content & Creative","Product + lifestyle photography","Sales-grade catalog & ad creative","Service","P0","Hybrid","In-house + freelance","60000","20000","M1"),
("Group","Content & Creative","Short-form video / reels engine","3-5 reels/wk per brand; organic + ads","Service","P0","In-house: Arihant","Content pod + CapCut Pro","20000","15000","M1"),
("Group","Content & Creative","Design subscription","Unlimited on-brand creatives at speed","In-house","P0","In-house: Designomics","Designomics + Adobe CC / Canva Teams","0","6000","M1"),
("Group","Content & Creative","Long-form video repurposing","Turn one shoot into 20 clips (Opus/Descript)","SaaS","P1","In-house: Arihant","Opus Clip / Descript","0","4000","M2"),
("Aurra Hype","Content & Creative","Drop lookbooks + campaign shoots","Hype cycle around each drop","Service","P1","Hybrid","Designomics + models","40000","25000","M2"),
# AI & Creative production
("Group","AI & Creative","AI UGC ad-video studio","Product URL → ready UGC ad (Hermes Agent)","SaaS","P1","In-house: Arihant","Higgsfield","0","7000","M2"),
("Group","AI & Creative","AI cinematic video (hero/brand films)","Veo/Runway/Kling for premium spots & drops","SaaS","P1","In-house: Arihant","Google Flow (Veo 3.1) / Runway / Kling","0","9000","M2"),
("Group","AI & Creative","AI image / product visuals","On-model + concept imagery without shoots","SaaS","P0","In-house: Designomics","Midjourney + Krea + Ideogram + Photoroom","0","6000","M1"),
("Group","AI & Creative","AI voice + avatar / spokesperson","Voiceovers + talking-avatar promos in many languages","SaaS","P2","In-house: Arihant","ElevenLabs + HeyGen","0","5000","M3"),
("Group","AI & Creative","AI copy + creative brains","Copy, scripts, descriptions, briefs","SaaS","P0","In-house: Arihant","ChatGPT Team + Claude","0","6000","M1"),
# Performance Marketing
("Aurra Hype","Performance Marketing","Meta Ads — always-on","Primary D2C revenue driver","Service+Spend","P0","In-house: Arihant","Meta Ads Manager","0","150000","M1"),
("Aurra Hype","Performance Marketing","Google Ads (Search+Shopping+PMax)","High-intent + Shopping feed","Service+Spend","P1","In-house: Arihant","Google Ads + Merchant Center","0","80000","M2"),
("Designomics","Performance Marketing","Meta + Google (gifting + corporate)","Seasonal + corporate intent","Service+Spend","P1","In-house: Arihant","Meta + Google","0","80000","M2"),
("Loop In","Performance Marketing","Local + LinkedIn lead-gen","B2B events + weddings/corporate","Service+Spend","P1","In-house: Arihant","Google Local + LinkedIn Ads","0","40000","M2"),
("Group","Performance Marketing","Creative analytics (winner-finding)","Element-level insight; lower CAC","SaaS","P1","In-house: Arihant","Motion / Superads","0","9000","M2"),
("Aurra Hype","Performance Marketing","Profit analytics + attribution","True profit after ads/COGS/RTO per creative","SaaS","P1","Buy","Triple Whale / Lifetimely","0","8000","M2"),
("Group","Performance Marketing","TikTok / YT Shorts / Pinterest ads","Diversify channels beyond Meta","Service+Spend","P2","In-house: Arihant","TikTok/YouTube/Pinterest Ads","0","40000","M3"),
# CRM & Retention
("Aurra Hype","CRM & Retention","Email + SMS lifecycle flows","Cart, win-back, post-purchase = repeat revenue","SaaS","P0","In-house: Arihant","Klaviyo","0","9000","M1"),
("Group","CRM & Retention","WhatsApp Business API + broadcast","#1 India channel: updates, drops, offers","SaaS","P0","Buy","AiSensy / WATI / Interakt","5000","6000","M1"),
("Group","CRM & Retention","Loyalty + referral program","Repeat purchase + WOM CAC cut","SaaS","P1","Buy","Zithara / Loox referrals","0","4000","M2"),
("Group","CRM & Retention","Advanced SMS marketing","Conversational SMS at scale (global best-practice)","SaaS","P2","Buy","Attentive / Postscript-style","0","6000","M3"),
("Group","CRM & Retention","Sales CRM (B2B pipeline)","Events + corporate gifting leads → deals","SaaS","P1","Buy","HubSpot / Zoho CRM","0","5000","M2"),
("Group","CRM & Retention","LTV / cohort / churn analytics","Predict LTV, know when customers churn","SaaS","P2","Buy","Peel Insights / Lifetimely","0","6000","M3"),
# Chat & Social Automation
("Group","Chat & Social Automation","IG comment-to-DM + DM flows","Auto-convert comments/DMs to sales","SaaS","P1","In-house: Arihant","ManyChat","0","6000","M2"),
("Group","Chat & Social Automation","WhatsApp chat automation + bot","Orders, FAQs, catalog, checkout on WhatsApp","SaaS","P1","In-house: Arihant","AiSensy + n8n","0","4000","M2"),
("Group","Chat & Social Automation","Telegram bot / broadcast","Community + drops + VIP list","Automation","P2","In-house: Arihant","Telegram Bot API + n8n","0","500","M3"),
("Group","Chat & Social Automation","Facebook/Messenger automation","Lead capture + retargeting handoff","SaaS","P2","In-house: Arihant","ManyChat","0",0,"M3"),
("Group","Chat & Social Automation","AI support chatbot (site+chat)","24x7 sizing/returns/order status","SaaS","P1","In-house: Arihant","Gorgias AI / Chatbase","0","6000","M2"),
# AI & Automation backbone
("Group","AI & Automation","Automation backbone + AI agents","Connect store, ads, CRM, WhatsApp, sheets","SaaS","P0","In-house: Arihant","n8n (self-host) + Make","6000","3000","M1"),
("Group","AI & Automation","Claude + MCP command layer","Query dashboards/DBs, run agents via MCP connectors","SaaS","P1","In-house: Arihant","Claude + MCP servers","0","4000","M2"),
("Group","AI & Automation","Gemini / multimodal assist","Docs, research, image+video understanding","SaaS","P1","In-house: Arihant","Google Gemini","0","2000","M2"),
("Group","AI & Automation","No-code AI agent builder","Ops agents (research, outreach, ops) fast","SaaS","P2","In-house: Arihant","Gumloop / Lindy / Relay","0","4000","M3"),
("Group","AI & Automation","Ops dashboards + daily digest","Auto WhatsApp digest: sales/ads/stock","Automation","P1","In-house: Arihant","n8n + Sheets + GPT","0","1000","M2"),
# Sourcing & Manufacturing
("Aurra Hype","Sourcing & Manufacturing","Apparel manufacturing partners","Reliable blanks + cut-and-sew","Service","P0","Buy","Tiruppur / Ludhiana / Bengaluru","0",0,"M1"),
("Aurra Hype","Sourcing & Manufacturing","DTG / DTF print partner","Small-batch drops, no big MOQ","Service","P0","Hybrid","Designomics + Qikink","0",0,"M1"),
("Designomics","Sourcing & Manufacturing","Merch & gifting supplier network","Mugs, bottles, candles, kits at scale","Service","P0","Buy","IndiaMART verified / agents","0",0,"M1"),
("Group","Sourcing & Manufacturing","Print-on-demand fallback","Zero-inventory long-tail SKUs","Service","P1","Buy","Qikink / Printrove","0",0,"M2"),
("Aurra Hype","Sourcing & Manufacturing","QC + tech-pack system","Consistent sizing/quality; fewer returns","Service","P1","In-house: Designomics","Tech packs + QC checklist","10000",0,"M2"),
("Group","Sourcing & Manufacturing","Global sourcing / import option","Trims, accessories, specialty items","Service","P3","Buy","Alibaba / verified importers","0",0,"M3"),
# Packaging & Unboxing
("Group","Packaging & Unboxing","Custom branded packaging","Unboxing = free marketing + value","Service","P0","In-house: Designomics","Designomics + Bizongo / Packman","35000","25000","M1"),
("Designomics","Packaging & Unboxing","Custom gift-box engine","Core product: build-your-own box UX","Service","P0","In-house: Designomics","Designomics + configurator","20000",0,"M1"),
("Group","Packaging & Unboxing","Inserts, cards, stickers, QR","Reviews, referrals, repeat-order QR","Service","P1","In-house: Designomics","Designomics print","0","8000","M2"),
("Group","Packaging & Unboxing","Sustainable / premium packaging line","Higher perceived value + ESG story","Service","P2","In-house: Designomics","Recyclable/kraft + custom","15000","10000","M3"),
# Inventory & Warehouse
("Group","Inventory & Warehouse","Inventory management system","Single stock source across channels","SaaS","P0","Buy","Zoho Inventory / Unicommerce","0","5000","M1"),
("Group","Inventory & Warehouse","Barcode / SKU + labels","Scan picking; stops overselling","SaaS","P1","Buy","Zoho + scanner","8000","1500","M2"),
("Group","Inventory & Warehouse","Warehouse / stockroom + WMS-lite","Organised bins for store+online","Service","P1","Buy","Racking + Zoho","50000",0,"M2"),
("Group","Inventory & Warehouse","Multi-channel stock sync","Auto-sync to Amazon/Flipkart/Q-comm","SaaS","P1","Buy","Unicommerce / Vinculum","0","8000","M2"),
("Group","Inventory & Warehouse","Demand forecasting","Predict winners; avoid dead stock","SaaS","P3","In-house: Arihant","Inventory data + GPT model","0","2000","M3"),
# Order & Logistics
("Group","Order & Logistics","Shipping aggregator","Cheapest courier per pincode + NDR","SaaS","P0","Buy","Shiprocket / iThink","0","4000","M1"),
("Group","Order & Logistics","COD + RTO control","Cut RTO losses (apparel killer)","SaaS","P1","Buy","GoKwik / Shiprocket","0","5000","M2"),
("Group","Order & Logistics","Order management system (OMS)","Unified multi-channel order queue","SaaS","P1","Buy","Unicommerce","0","6000","M2"),
("Group","Order & Logistics","1-click / UPI fast checkout","Cut checkout drop-off","SaaS","P1","Buy","GoKwik / Shopflo / Razorpay Magic","0","5000","M2"),
("Group","Order & Logistics","Returns / exchange automation","Streamlined apparel returns; protect margin","SaaS","P1","Buy","Return Prime / Loop-style","0","3000","M2"),
("Aurra Hype","Order & Logistics","Hyperlocal same-day (Bengaluru)","Same-day for local + store","Service","P2","Buy","Porter / Dunzo / Shiprocket Quick","0","4000","M3"),
# Quick Commerce & Marketplaces
("Aurra Hype","Q-Commerce & Marketplaces","Amazon + Flipkart listings","Discovery + trust; A+ content","Service","P1","In-house: Arihant","Seller Central + Flipkart","0","10000","M2"),
("Aurra Hype","Q-Commerce & Marketplaces","Myntra / Ajio (fashion)","Category-native fashion buyers","Service","P2","Buy","Myntra / Ajio onboarding","0",0,"M3"),
("Designomics","Q-Commerce & Marketplaces","Meesho (value gifting volume)","0% commission, high-volume SKUs","Service","P2","In-house: Arihant","Meesho","0",0,"M3"),
("Designomics","Q-Commerce & Marketplaces","Q-commerce pilot (gifting SKUs)","Impulse gifting via Blinkit/Zepto/Instamart","Service+Fee","P3","Buy","Blinkit/Zepto/Instamart","100000","250000","M3"),
("Group","Q-Commerce & Marketplaces","Marketplace ads","Sponsored products; win visibility","Service+Spend","P2","In-house: Arihant","Amazon/Flipkart Ads","0","40000","M3"),
("Designomics","Q-Commerce & Marketplaces","Global handmade marketplace","Etsy for stickers/gifting (export optional)","Service","P3","In-house: Arihant","Etsy","0","2000","M3"),
# Influencer & PR
("Aurra Hype","Influencer & PR","Influencer seeding (micro+mid)","Hype & reach for streetwear","Service+Spend","P1","Buy","Local + GRIN / Aspire","0","100000","M2"),
("Group","Influencer & PR","Influencer/creator CRM platform","Manage seeding, affiliates, payouts at scale","SaaS","P2","In-house: Arihant","GRIN / Aspire / Upfluence","0","15000","M3"),
("Group","Influencer & PR","Affiliate program","Performance-based reach","SaaS","P2","In-house: Arihant","Refersion / GoAffPro","0","3000","M3"),
("Group","Influencer & PR","PR + media outreach","Press features = trust + SEO","Service","P2","In-house: Arihant","HARO/Connectively + local PR","0","20000","M3"),
# Retail / Offline
("Aurra Hype","Retail / Offline","Retail POS + store ops","Unify store + online stock & billing","SaaS","P0","Buy","Shopify POS / GoFrugal","10000","3000","M1"),
("Aurra Hype","Retail / Offline","Store launch (12 Jul) amplification","Turn launch into content + PR + footfall","Service","P0","In-house: Arihant","Arihant + Loop In","50000",0,"M1"),
("Group","Retail / Offline","Pop-ups / exhibitions / activations","Direct sales + UGC + email capture","Service","P2","Hybrid","Loop In + Designomics","40000","30000","M3"),
("Group","Retail / Offline","Offline ads (OOH / print / local)","Local awareness around store & events","Service+Spend","P3","In-house: Arihant","Local OOH / print","0","30000","M3"),
# Finance & Compliance
("Group","Finance & Compliance","Accounting + GST software","Clean books, GST across entities","SaaS","P0","Buy","Zoho Books / TallyPrime","6000","2500","M1"),
("Group","Finance & Compliance","Payment gateway + reconciliation","UPI/cards/COD; auto-reconcile","SaaS","P0","Buy","Razorpay / Cashfree","0",0,"M1"),
("Group","Finance & Compliance","CA + bookkeeping + GST filing","Compliance; investor-ready books","Service","P0","Buy","CA firm (retainer)","0","15000","M1"),
("Group","Finance & Compliance","Entity + structure setup","LLP/Pvt Ltd, GST, IEC for exports","Service","P0","Buy","CA + CS","40000",0,"M1"),
("Group","Finance & Compliance","Spend / corporate cards + expense","Control burn across 4 brands","SaaS","P1","Buy","Volopay / Jify / RazorpayX","0","2000","M2"),
("Group","Finance & Compliance","Unit-economics + P&L dashboard","CAC, contribution margin, RTO per brand","Service","P1","In-house: Arihant","Sheets/Looker + Arihant","0","2000","M2"),
("Aurra Hype","Finance & Compliance","Revenue-based financing (inventory)","Fund inventory/ads without dilution","Service","P3","Buy","Wayflyer / Velocity-style","0",0,"M3"),
# Legal & IP
("Group","Legal & IP","Trademark portfolio management","Own names/logos; renewals","Service","P0","Buy","IP attorney retainer","0","6000","M1"),
("Group","Legal & IP","Contracts + policies + agreements","Supplier, influencer, employment, T&Cs","Service","P1","Buy","Legal retainer / Vakilsearch","0","6000","M2"),
("Group","Legal & IP","Data privacy / DPDP compliance","India DPDP Act + global if exporting","Service","P2","Buy","Legal + consent tooling","0","3000","M3"),
# HR & Team Ops
("Group","HR & Team Ops","Core hiring (mktg, ops, design)","People to run the machine","Service","P1","Hybrid","Naukri/LinkedIn + referrals","0",0,"M2"),
("Group","HR & Team Ops","Payroll + attendance + compliance","Compliant salaries across brands","SaaS","P1","Buy","RazorpayX Payroll / Keka","0","3000","M2"),
("Group","HR & Team Ops","SOPs + training library","Repeatable ops; fast onboarding","Service","P1","In-house: Arihant","Notion + Loom","0","1500","M2"),
("Group","HR & Team Ops","Contractor / freelancer management","Designers, editors, creators payouts","SaaS","P2","In-house: Arihant","Notion + Razorpay payouts","0","1000","M3"),
# Project & Internal Ops
("Group","Project & Internal Ops","Project / task management","One system across 4 brands","SaaS","P0","In-house: Arihant","Notion / ClickUp","0","3000","M1"),
("Group","Project & Internal Ops","AI calendar + focus/time","Founder & team time management","SaaS","P1","In-house: Arihant","Motion (usemotion) / Reclaim","0","3000","M2"),
("Group","Project & Internal Ops","Meeting notes AI","Auto notes + action items from calls","SaaS","P1","In-house: Arihant","Fathom / Fireflies","0","2000","M2"),
("Group","Project & Internal Ops","Central docs + asset library (DAM)","Brand assets, contracts, SOPs in one place","SaaS","P1","In-house: Arihant","Google Drive / Notion","0","1500","M2"),
("Loop In","Project & Internal Ops","Event project + vendor mgmt","Timelines, vendors, checklists per event","SaaS","P1","Buy","Monday.com + Aisle Planner","0","4000","M2"),
# Analytics & BI
("Group","Analytics & BI","Web + funnel analytics","GA4 + server-side + attribution","SaaS","P0","In-house: Arihant","GA4 + Meta CAPI + Stape","5000","2000","M1"),
("Group","Analytics & BI","Group KPI dashboard","One board: rev, CAC, AOV, ROAS per brand","Service","P1","In-house: Arihant","Looker Studio + Arihant","0","2000","M2"),
("Aurra Hype","Analytics & BI","Session replay + heatmaps","See where users drop; fix UX","SaaS","P2","In-house: Arihant","Microsoft Clarity (free) / Hotjar","0",0,"M3"),
]
aligns=["center","left","left","left","left","center","center","left","left","center","center","center"]
r=5
for i,row in enumerate(S,1):
    biz,dom,item,what,typ,pri,inh,ven,setup,mon,ph=row
    vals=[i,biz,dom,item,what,typ,pri,inh,ven,int(setup),(int(mon) if str(mon).lstrip('-').isdigit() else mon),ph]
    for ci,v in enumerate(vals,1):
        cell=ms.cell(row=r,column=ci,value=v); cell.border=border
        cell.font=Font(name=FONT,size=9,color="222222")
        cell.alignment=Alignment(vertical="top",wrap_text=True,horizontal=aligns[ci-1])
    ms.cell(row=r,column=2).fill=PatternFill("solid",fgColor=BIZC.get(biz,"FFFFFF"))
    pc=ms.cell(row=r,column=7); pc.font=Font(name=FONT,size=9,bold=True,color="FFFFFF"); pc.fill=PatternFill("solid",fgColor=PRI.get(pri,"999999"))
    ms.cell(row=r,column=10).number_format=RUPEE; ms.cell(row=r,column=11).number_format=RUPEE
    if r%2==0:
        for ci in [1,3,4,5,6,8,9,12]: ms.cell(row=r,column=ci).fill=PatternFill("solid",fgColor="FAFBFC")
    r+=1
tot=r
ms.cell(row=tot,column=9,value="TOTALS →").font=Font(name=FONT,bold=True,size=10,color=NAVY)
ms.cell(row=tot,column=9).alignment=Alignment(horizontal="right")
ms.cell(row=tot,column=10,value=f"=SUM(J5:J{tot-1})").number_format=RUPEE
ms.cell(row=tot,column=11,value=f"=SUM(K5:K{tot-1})").number_format=RUPEE
for ci in (10,11):
    ms.cell(row=tot,column=ci).font=Font(name=FONT,bold=True,size=10,color=NAVY); ms.cell(row=tot,column=ci).fill=PatternFill("solid",fgColor=YELLOW)
ms.freeze_panes="A5"; ms.auto_filter.ref=f"A4:L{tot-1}"
widths(ms,[4,12,20,28,42,11,8,19,25,11,12,7])

print("master stack rows:", len(S))
wb.save("/sessions/bold-epic-ramanujan/mnt/outputs/Harshdeep_Group_Master_Stack_FULL.xlsx")
print("saved base + master")
