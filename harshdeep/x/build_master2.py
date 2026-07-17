# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F2A44"; BLUE="2E5AAC"; TEAL="0E7C7B"; ORANGE="C05621"; PURPLE="5B3A8C"; PLUM="7A2E5B"
LT_GREY="F4F5F7"; YELLOW="FFF3C4"; RED="B02A37"
FONT="Calibri"; RUPEE='"₹"#,##0'
thin=Side(style="thin",color="C9CED6"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
PRI={"P0":RED,"P1":ORANGE,"P2":TEAL,"P3":"7A8288"}
P="/sessions/bold-epic-ramanujan/mnt/outputs/Harshdeep_Group_Master_Stack_FULL.xlsx"
wb=load_workbook(P)

def sheet(name):
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False; return ws
def banner(ws,title,sub,ncols,color=NAVY):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
    c=ws.cell(row=1,column=1,value=title); c.font=Font(name=FONT,bold=True,size=15,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor=color); c.alignment=Alignment(horizontal="left",vertical="center",indent=1); ws.row_dimensions[1].height=28
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncols)
    c=ws.cell(row=2,column=1,value=sub); c.font=Font(name=FONT,italic=True,size=9,color="333333")
    c.fill=PatternFill("solid",fgColor=LT_GREY); c.alignment=Alignment(horizontal="left",indent=1); ws.row_dimensions[2].height=16
def hdr(ws,row,headers,color=NAVY):
    for i,h in enumerate(headers,1): ws.cell(row=row,column=i,value=h)
    for cc in range(1,len(headers)+1):
        cell=ws.cell(row=row,column=cc); cell.font=Font(name=FONT,bold=True,color="FFFFFF",size=9)
        cell.fill=PatternFill("solid",fgColor=color); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
def rows_out(ws,start,rows,aligns,fontsize=9,pri_col=None):
    r=start
    for row in rows:
        for ci,v in enumerate(row,1):
            cell=ws.cell(row=r,column=ci,value=v); cell.border=border
            cell.font=Font(name=FONT,size=fontsize,color="222222"); cell.alignment=Alignment(vertical="top",wrap_text=True,horizontal=aligns[ci-1])
        if pri_col:
            pv=row[pri_col-1]; pc=ws.cell(row=r,column=pri_col)
            pc.font=Font(name=FONT,size=fontsize,bold=True,color="FFFFFF"); pc.fill=PatternFill("solid",fgColor=PRI.get(pv,"999999"))
        if r%2==0:
            for ci in range(1,len(row)+1):
                if pri_col and ci==pri_col: continue
                ws.cell(row=r,column=ci).fill=PatternFill("solid",fgColor="FAFBFC")
        r+=1
    return r
def widths(ws,ws_widths):
    for i,w in enumerate(ws_widths,1): ws.column_dimensions[get_column_letter(i)].width=w

# ---------- BRAND TAB HELPER ----------
def brand_tab(name, color, headline, positioning, stack, wcols):
    ws=sheet(name)
    banner(ws,name+" — Brand Playbook",headline,5,color)
    ws.merge_cells(start_row=4,start_column=1,end_row=4,end_column=5)
    c=ws.cell(row=4,column=1,value="POSITIONING:  "+positioning); c.font=Font(name=FONT,size=10,italic=True,color="333333")
    c.alignment=Alignment(wrap_text=True,vertical="top"); ws.row_dimensions[4].height=44
    hdr(ws,6,["Priority","Workstream","Key action / what to build","Tool / Owner","Proof / outcome"])
    end=rows_out(ws,7,stack,["center","left","left","left","left"],pri_col=1)
    ws.freeze_panes="A7"; widths(ws,wcols)
    return ws

# AURRA HYPE
brand_tab("Aurra Hype", "C05621",
 "D2C streetwear / hype apparel · Bengaluru · store+office launch 12 Jul",
 "The drop-driven Indian streetwear label. Win = desirable product + hype/scarcity rhythm + disciplined performance marketing with healthy unit economics. Copy the global streetwear playbook (limited drops, community, resale hype).",
 [
 ("P0","Brand & Trademark","Refine identity vs crowded 'Aura' space; file trademark","Designomics + IP attorney","Distinct, owned brand"),
 ("P0","Store","Shopify build, reviews, upsell, 1-tap UPI/COD checkout","Arihant + GoKwik","CVR > 2%, fast store"),
 ("P0","Retail","POS unifying store+online stock; amplify 12 Jul launch as content/PR","Arihant + Loop In","500+ footfall, viral launch"),
 ("P0","Creative","Catalog + lifestyle shoot; reels engine 3-5/wk","Content pod","90+ assets in M1"),
 ("P0","Paid","Meta always-on; pixel+CAPI; creative testing","Arihant + Motion","ROAS baseline set"),
 ("P1","Paid scale","Add Google Shopping/PMax; scale winners; profit analytics","Arihant + Triple Whale","ROAS >= 2.5"),
 ("P1","AI creative","Higgsfield UGC ads + Runway/Veo hero drop films","Arihant","Cheaper, faster ad volume"),
 ("P1","Retention","Klaviyo email/SMS + WhatsApp flows; loyalty","Arihant + AiSensy","Repeat-rate rising"),
 ("P1","Drops","Drop calendar w/ scarcity, waitlists, early-access","Arihant","Sell-through per drop"),
 ("P1","Influencer","Micro+mid seeding; UGC pipeline","GRIN/Aspire + local","Reach + UGC library"),
 ("P1","Marketplace","Amazon + Flipkart w/ A+ content","Arihant","Live on 2 marketplaces"),
 ("P2","Marketplace","Myntra/Ajio fashion onboarding","Buy","Category-native sales"),
 ("P2","Logistics","RTO/COD control; same-day Bengaluru","GoKwik + Porter","RTO trending down"),
 ("P3","Community","Drop community (WhatsApp/Telegram VIP + Discord)","Arihant","Owned hype channel"),
 ],[9,16,40,22,26])

# DESIGNOMICS
brand_tab("Designomics", "1E7D34",
 "Personalised gifting + merch + apparel · corporate gifting · in-house design/print arm",
 "Two demand engines: impulse/occasion D2C gifting AND contract corporate gifting. Highest natural margin of the group; premium comes from a great customization UX. Also produces the group's design, print and packaging.",
 [
 ("P0","Replatform","Move off Wix; custom gift-box configurator; apps + speed","Arihant + Kickflip","Higher AOV, better UX"),
 ("P0","Supply","Merch/gifting supplier network + POD fallback","Buy + Qikink","Reliable catalog"),
 ("P0","Packaging","Own the group's packaging + unboxing system","Designomics","Branded unboxing on every order"),
 ("P1","Corporate B2B","Outbound to companies; catalog; quote-to-order; GST bulk","Founder + HubSpot","5+ corporate deals in pipeline"),
 ("P1","Paid","Meta + Google on gifting & corporate intent; festive push","Arihant","Profitable seasonal spend"),
 ("P1","AI creative","AI product visuals (Midjourney/Krea) for endless SKUs","Designomics","Catalog-ready visuals cheaply"),
 ("P1","Retention","Occasion automation (birthday/festival/anniversary)","Klaviyo + AiSensy","Recurring gifting revenue"),
 ("P2","Marketplace","Meesho (volume) + Amazon gifting + Etsy (export optional)","Arihant","New discovery channels"),
 ("P2","Workshops","Productise workshops (DTF/patch kits) as events + content","Loop In + Designomics","Community + revenue"),
 ("P3","Q-commerce","Pilot Blinkit/Zepto on high-margin fast SKUs only","Buy","Impulse gifting channel"),
 ],[9,16,40,22,26])

# LOOP IN
brand_tab("Loop In (rebrand)", "5B3A8C",
 "Event management & experiential · REBRAND FIRST (Canada name/domain clash) · group's live-activation arm",
 "Runs the group's launches, pop-ups, drops and corporate-gifting experiences, and sells events externally. MUST rebrand: 'Loop In Events' + loopinevents.com belong to a Canadian planner. Secure a distinct, trademarked name + domain before scaling.",
 [
 ("P0","REBRAND","IP-India search (Class 41 & 35); pick distinct name; lock domain+handles; file TM","Founder + IP attorney","Owned, trademarkable brand"),
 ("P0","Web","Lead-gen site + portfolio + packages + booking","Arihant","Enquiries flowing"),
 ("P1","Pipeline","Local + LinkedIn lead-gen; sales CRM; quote system","Arihant + HubSpot","3 events booked"),
 ("P1","Ops","Event project + vendor management system","Monday.com + Aisle Planner","On-time, on-budget events"),
 ("P1","Cross-sell","Run Aurra Hype launches + Designomics corporate experiences","Group","Internal revenue + case studies"),
 ("P2","Content","Film every event as reels/case studies/testimonials","Content pod","Portfolio + social proof"),
 ("P2","Automation","Post-event feedback + testimonial + upsell flow","Arihant + n8n","Reviews + repeat bookings"),
 ("P3","Scale","Signature owned event/IP (e.g. streetwear x music night)","Group","Brand-building flagship"),
 ],[9,16,40,22,26])

# ARIHANT
brand_tab("Arihant Digital", "2E5AAC",
 "~10-yr digital agency = the in-house engine AND a productised external offer",
 "The unfair advantage. Builds & markets the 3 brands at internal cost, then sells that battle-tested, AI-native capability to outside clients. The 3 brands become its case studies. This is how the group funds itself and compounds.",
 [
 ("P0","Internal delivery","Deliver web, ads, SEO/GEO, automation, analytics to all 3 brands","Arihant team","3 brands live & scaling"),
 ("P1","Productise","Package offers: 'AI-native growth', 'store build', 'automation setup'","Founder","Clear sellable packages"),
 ("P1","Proof assets","Turn brand wins into case studies + before/after numbers","Arihant","Portfolio of real results"),
 ("P1","AI edge","Lead with AI creative (Higgsfield) + automation (n8n/MCP) as differentiator","Arihant","Premium positioning"),
 ("P2","Outbound","Cold + inbound + referrals; sell to D2C brands like your own","Founder + HubSpot","First external clients"),
 ("P2","Retainers","Move clients to monthly retainers (predictable revenue)","Arihant","MRR base"),
 ("P3","Scale team","Hire pod-based delivery; SOPs; margins","Founder","Scalable agency"),
 ],[9,16,40,22,26])

# ---------- AI & CREATIVE TOOLS ----------
ai=sheet("AI & Creative Stack")
banner(ai,"AI & CREATIVE STACK (2026, world-class)","The AI engine world-class brands use. USD pricing indicative — verify at signup; models move fast.",6,PLUM)
hdr(ai,4,["Category","Tool","What it does","Best for","Approx price (USD/mo)","Priority"])
A=[
("Video — UGC ads","Higgsfield (Hermes Agent)","Product URL → ready UGC/CGI ad video; aggregates 15+ models","Meta/TikTok ad volume, testimonials, unboxings","$15–84","P1"),
("Video — cinematic","Google Flow (Veo 3.1)","Best overall AI video + native audio, 4K","Hero brand films, drop teasers","~$20+","P1"),
("Video — control","Runway (Gen-4.5)","Keyframes, motion brush, character consistency, editor","Marketer-controlled brand video","$15–95","P1"),
("Video — value","Kling 3.0","Cheapest premium model (~$0.10/s), multi-shot","High-volume iteration","pay-as-you-go","P2"),
("Video — repurpose","Opus Clip / Descript","1 long video → many captioned clips; edit by text","Reels/Shorts factory, podcasts","$15–30","P1"),
("Image — concept","Midjourney","Best-in-class stylised imagery","Moodboards, campaign concepts","$10–60","P0"),
("Image — realtime","Krea / Leonardo","Real-time + model hub (Flux etc.)","Fast product/lifestyle variations","$10–60","P1"),
("Image — text/logo","Ideogram","Best text-in-image + logo/poster","Posters, packaging art, typography","$8–48","P2"),
("Image — product","Photoroom / Flair","Background removal + AI product scenes","Catalog + on-model shots","$10–40","P0"),
("Voice","ElevenLabs","Realistic AI voiceover, many languages","Ad VO, IVR, dubbing","$5–99","P2"),
("Avatar","HeyGen","Talking-avatar spokesperson videos","Multilingual promos, explainers","$29–89","P2"),
("Copy / brains","ChatGPT Team + Claude","Copy, scripts, briefs, research, agents","All content + strategy + coding","$25–30/seat","P0"),
("Multimodal","Google Gemini","Docs, deep research, image+video understanding","Research, analysis, Workspace","$20+","P1"),
("AI agents / MCP","Claude + MCP servers","Connect AI to store/CRM/DBs; run agentic ops","Founder command layer, automations","usage","P1"),
("No-code agents","Gumloop / Lindy / Relay","Build ops agents (research, outreach, ops)","Sales/ops automation w/o code","$20–99","P2"),
("Automation","n8n (self-host) + Make","Wire every app + AI together","Backbone of all automations","$0–50","P0"),
("Chat automation","ManyChat","IG/FB/WhatsApp/Telegram/SMS comment+DM flows","Social selling automation","$15–139","P1"),
("Creative analytics","Motion","Auto-tags creative + performance; winner detection","Scaling ads profitably","$99–499","P1"),
("Profit analytics","Triple Whale","True profit/attribution per creative/channel","D2C financial truth","$129–499","P1"),
]
rows_out(ai,5,A,["left","left","left","left","center","center"],pri_col=6)
ai.freeze_panes="A5"; widths(ai,[18,22,40,30,16,9])

# ---------- AUTOMATION CATALOGUE ----------
au=sheet("Automation Catalogue")
banner(au,"AUTOMATION CATALOGUE","Build on n8n/Make + WhatsApp API + ManyChat + GPT/Claude(MCP). Channel-tagged.",7,TEAL)
hdr(au,4,["#","Automation","Channel","Business","Trigger → Action","Tools","Impact"])
AU=[
("Abandoned-cart recovery","WhatsApp+Email","Aurra/Design","Cart not completed → timed WA+email w/ product+offer","Shopify+Klaviyo+AiSensy+n8n","Recovers 8-15% carts"),
("Order status updates","WhatsApp","All","Placed/shipped/OFD/delivered → auto WA w/ tracking","Shiprocket+WA API+n8n","-60% WISMO tickets"),
("IG comment-to-DM funnel","Instagram","Aurra/Design","Comment keyword → auto DM link/coupon → capture","ManyChat","Turns engagement into sales"),
("IG/DM auto-reply + qualify","Instagram","All","DM in → AI reply, sizing/FAQ, qualify, handoff","ManyChat + Chatbase","24x7 IG selling"),
("WhatsApp catalog + checkout","WhatsApp","Aurra/Design","Browse catalog, order, pay on WhatsApp","AiSensy + Razorpay","High-converting channel"),
("Telegram drops/VIP broadcast","Telegram","Aurra","New drop → broadcast to VIP list + early access","Telegram Bot + n8n","Owned hype channel"),
("Facebook lead → CRM","Facebook","Loop In/Design","Lead ad → CRM → auto WA/email + assign","ManyChat + HubSpot + n8n","No lead dropped"),
("AI support chatbot","Site+WA","All","Any query → sizing/returns/status; human handoff","Gorgias AI / Chatbase","24x7 support"),
("Review + UGC request","WhatsApp+Email","Aurra/Design","Delivered +5d → ask review+photo, reward coupon","Judge.me/Loox+WA+n8n","Social proof + UGC"),
("AI ad-creative pipeline","Internal","All","New SKU → Higgsfield UGC ad + captions auto-drafted","Higgsfield+GPT+n8n","Ad volume, -70% time"),
("AI product imagery","Internal","Aurra/Design","New SKU → on-model + scene images auto-generated","Photoroom/Krea+n8n","Catalog-ready fast"),
("Daily founder digest","WhatsApp/Telegram","Group","8am → sales, spend, ROAS, top SKUs, low stock","Sheets+GA4+Meta+GPT+n8n","One-glance control"),
("Low-stock + auto-PO","WhatsApp","All","Stock<threshold → alert + draft PO to supplier","Zoho Inv+n8n","No stockouts on winners"),
("RTO/COD risk flag","WhatsApp","All","COD order → risk score → verify risky via WA","GoKwik/Shiprocket+n8n","Protects margin"),
("Lead → CRM → nurture","Multi","Loop In/Design","Any lead → CRM, intro+booking link, drip","Forms+HubSpot+n8n","Faster follow-up"),
("Occasion gifting reminder","WhatsApp+Email","Designomics","Birthday/festival nears → nudge + pre-filled cart","CRM+Klaviyo+WA","Recurring revenue"),
("Post-event feedback+upsell","WhatsApp+Email","Loop In","Event ends → feedback+testimonial+next offer","Forms+CRM+WA","Reviews + rebookings"),
("Influencer outreach at scale","IG/Email","Aurra","Weekly → find+DM micro-creators, track replies","Sheets+IG+GPT+n8n","Scales seeding cheaply"),
("Marketplace listing sync","Internal","Aurra","New/edited SKU → push to Amazon/Flipkart w/ A+ copy","Unicommerce+GPT+n8n","List everywhere fast"),
("Client-ready weekly report","Internal","Arihant","Mondays → branded KPI PDF auto-built","Looker+GPT+n8n","Sellable Arihant service"),
("Social scheduling + repurpose","Multi","Group","New asset → auto-schedule across IG/FB/YT/LinkedIn","Metricool+Opus+n8n","Consistent omni-presence"),
("Meeting → action items","Internal","Group","Call ends → AI notes + tasks to project tool","Fathom/Fireflies+Notion","Nothing falls through"),
]
r=5
for i,(nm,ch,biz,ta,tl,imp) in enumerate(AU,1):
    for ci,v in enumerate([i,nm,ch,biz,ta,tl,imp],1):
        cell=au.cell(row=r,column=ci,value=v); cell.border=border
        cell.font=Font(name=FONT,size=9,color="222222"); cell.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if ci==1 else "left"))
    if r%2==0:
        for ci in range(1,8): au.cell(row=r,column=ci).fill=PatternFill("solid",fgColor="FAFBFC")
    r+=1
au.freeze_panes="A5"; widths(au,[4,26,16,14,42,26,26])

# ---------- FOUNDER OS ----------
fo=sheet("Founder OS")
banner(fo,"FOUNDER OPERATING SYSTEM","One command center to run all 4 brands. Full detail in the Founder OS playbook doc.",4,NAVY)
hdr(fo,4,["Layer","What it is","Tools","Cadence / notes"])
FO=[
("1 · Command Center","Single dashboard: revenue, CAC, ROAS, AOV, cash, stock, pipeline — all 4 brands","Looker Studio + Triple Whale + Sheets (fed by n8n)","Live; check each morning"),
("2 · Daily digest","8am WhatsApp/Telegram summary + anomalies + to-dos","n8n + GPT + WhatsApp/Telegram","Daily, automated"),
("3 · AI Chief of Staff","Ask questions across your data via Claude + MCP connectors (Shopify, ads, CRM, Sheets)","Claude + MCP servers","On-demand, conversational"),
("4 · Work system","All brand workstreams, tasks, SOPs in one place","Notion / ClickUp","Weekly review"),
("5 · Time & focus","AI auto-schedules deep work, meetings, priorities","Motion (usemotion) / Reclaim","Daily auto-planning"),
("6 · Meetings","Auto notes + action items pushed to tasks","Fathom / Fireflies","Every call"),
("7 · Comms hub","One inbox: email, IG, WhatsApp, chat","Gorgias / shared inbox + Superhuman","Batch 2x/day"),
("8 · Money view","Spend, burn, runway, approvals across brands","Volopay/Jify + Zoho Books","Weekly"),
("9 · Approvals","Creative, spend, hires flow to you for 1-tap approve","n8n + WhatsApp/Slack","Async"),
("10 · Weekly scorecard","Auto scorecard vs targets per brand + team accountability","Looker + n8n","Monday review"),
]
rows_out(fo,5,FO,["left","left","left","left"])
fo.freeze_panes="A5"; widths(fo,[20,40,28,30])

# ---------- BENCHMARKS ----------
bm=sheet("Benchmarks")
banner(bm,"WORLD-CLASS BENCHMARKS — what leading brands use, by function","Copy proven playbooks. Match your function to the best-in-class default.",4,ORANGE)
hdr(bm,4,["Function","World-class default stack","Example brands (global/India)","Your move"])
BM=[
("D2C platform","Shopify (→ Plus at scale)","Gymshark, Allbirds, boAt, Sugar","Shopify for Aurra + Designomics"),
("Email/SMS retention","Klaviyo (+ Attentive/Postscript SMS)","Vuori, Ridge, Mamaearth","Klaviyo + WhatsApp (India)"),
("Profit analytics","Triple Whale (50k+ brands)","Ridge, Obvi","Triple Whale for Aurra"),
("Creative analytics","Motion","Vuori, Keeps, Ridge","Motion once ad spend scales"),
("AI ad creative","Higgsfield + Runway + Veo","AI-native D2C brands 2026","Higgsfield for UGC ad volume"),
("Chat automation","ManyChat (IG/WA) ","Creators + D2C at scale","ManyChat + AiSensy"),
("Support","Gorgias (AI-first)","Steve Madden, Princess Polly","Gorgias AI across brands"),
("Influencer","GRIN / Aspire / Upfluence","SKIMS, Glossier","GRIN once seeding scales"),
("Reviews/UGC","Okendo / Loox / Judge.me","Premium D2C","Judge.me/Loox on Shopify"),
("Subscriptions","Recharge / Loop","Athletic Greens","Only if gifting subscription boxes"),
("Returns","Loop Returns / Return Prime","Fashion D2C","Return Prime (India)"),
("Analytics/attribution","GA4 + server-side + Triple Whale","All serious D2C","GA4+CAPI+Stape + TW"),
("Automation/agents","n8n / Make + Claude MCP","AI-native ops teams","n8n backbone + MCP layer"),
("Project/knowledge","Notion + Linear/ClickUp","Most startups","Notion for 4-brand ops"),
("Meeting AI","Fathom / Fireflies / Granola","Sales-led orgs","Fathom for all calls"),
("Time management","Motion (usemotion)","Founders/execs","Motion for founder focus"),
]
rows_out(bm,5,BM,["left","left","left","left"])
bm.freeze_panes="A5"; widths(bm,[22,32,30,30])

# ---------- 90-DAY ROADMAP ----------
rd=sheet("90-Day Roadmap")
banner(rd,"90-DAY SCALE ROADMAP","M1 Foundation · M2 Engine · M3 Scale. 'Proof' = milestone that shows it's working.",6)
hdr(rd,4,["Phase","Business","Workstream","Key actions","Owner","Proof / KPI"])
R=[
("M1 · Foundation","Group","Legal + Brand","Clean entities/GST; TM search+file x3; rebrand Loop In; unify brands","Founder+CA+Designomics","3 TMs filed · Loop In relaunched"),
("M1 · Foundation","Group","Tech backbone","Shopify+domains+WA API+payments+analytics+n8n+inventory+accounting live","Arihant","All core systems connected"),
("M1 · Foundation","Aurra Hype","Store + launch","Store upgrade, catalog shoot, POS, amplify 12 Jul launch","Arihant+Designomics+Loop In","500+ footfall · content spike"),
("M1 · Foundation","Group","Content + AI","Content pod live; AI creative stack (Higgsfield/Midjourney) set up","Arihant","90+ assets · AI pipeline live"),
("M1 · Foundation","Group","Packaging","Branded boxes/inserts ordered; gift-box configurator live","Designomics","Branded unboxing shipping"),
("M1 · Foundation","Aurra Hype","Paid on","Meta always-on; pixel/CAPI; first tests","Arihant","ROAS baseline set"),
("M2 · Engine","Aurra Hype","Scale paid + CRO","Add Google/PMax; scale winners; Klaviyo+WA flows","Arihant+Triple Whale","ROAS >= 2.5 · repeat flows live"),
("M2 · Engine","Group","Automations","Cart recovery, order updates, IG/WA bots, digest, reviews","Arihant","15+ automations live"),
("M2 · Engine","Aurra Hype","Marketplaces","Amazon + Flipkart w/ A+ content + ads","Arihant","Live on 2 marketplaces"),
("M2 · Engine","Designomics","Corporate gifting","Outbound + catalog + quote system; festive push","Founder+Arihant","5+ corporate deals in pipeline"),
("M2 · Engine","Loop In","Pipeline","Relaunch; lead-gen; CRM; 2-3 events booked","Loop In+Arihant","3 events booked · case studies"),
("M2 · Engine","Group","Founder OS","Command center + daily digest + AI chief-of-staff live","Arihant","One-pane control of 4 brands"),
("M3 · Scale","Aurra Hype","Drops + influencers","Drop calendar; seeding; scale winning ad budgets","Arihant+creators","Record revenue month"),
("M3 · Scale","Designomics","Q-comm + POD","Test Blinkit/Zepto; POD long-tail; festive campaigns","Founder+Arihant","Q-comm pilot live"),
("M3 · Scale","Group","Offline + PR","Pop-ups/exhibitions/activations; PR features","Loop In+Designomics","3 activations · press"),
("M3 · Scale","Arihant","Productise + sell","Package offer; land first external clients off proof","Founder+Arihant","First external clients (proof)"),
("M3 · Scale","Group","Prove numbers","Full unit-economics; investor/partner deck from real data","Arihant+CA","Documented growth curve"),
]
r=5; pc={"M1":"E8EEF9","M2":"E2EFDA","M3":"FCE4D6"}
for row in R:
    for ci,v in enumerate(row,1):
        cell=rd.cell(row=r,column=ci,value=v); cell.border=border
        cell.font=Font(name=FONT,size=9,color="222222",bold=(ci==1)); cell.alignment=Alignment(vertical="top",wrap_text=True,horizontal="left")
    rd.cell(row=r,column=1).fill=PatternFill("solid",fgColor=pc.get(row[0][:2],"FFFFFF"))
    r+=1
rd.freeze_panes="A5"; widths(rd,[16,13,18,44,22,32])

# ---------- MARKETPLACES ----------
qc=sheet("Q-Commerce & Marketplaces")
banner(qc,"Q-COMMERCE & MARKETPLACE CHANNELS","Real 2026 India costs & fit. Verify terms at signup.",7)
hdr(qc,4,["Channel","Best for","Onboarding cost / model","Commission","Fit","Priority","Notes"])
Q=[
("Own D2C (Shopify)","All — highest margin","Owned stack","~2% gateway only","Core","P0","Always the priority channel"),
("WhatsApp Commerce","Repeat + gifting + drops","API ~₹5-10k/mo","per-conversation","High","P0","India's highest-converting channel"),
("Amazon India","Apparel + gifting","Free acct; pay-per-sale; optional FBA","~10-20%+","High","P1","A+ content + Sponsored Products"),
("Flipkart","Apparel + gifting","Free listing; per-sale","~5-20%","High","P1","Big Billion Days spikes"),
("Meesho","Value gifting/stickers volume","Free; 0% commission","Low","Medium","P2","Great for low-price high-volume"),
("Myntra","Aurra fashion-native","Brand approval","~25-40% eff.","Med-High","P2","Strong for streetwear"),
("Ajio","Aurra fashion","Brand onboarding","high","Medium","P2","Reliance ecosystem"),
("Etsy (export)","Designomics stickers/gifting","Free; listing+txn fees","~6.5%+","Med (global)","P3","Optional international reach"),
("Blinkit","Impulse gifting SKUs","~₹25k/SKU/state (ad credits)+₹2-3L/mo","15-25%+","Med (need >65% margin)","P3","Selective; fast movers only"),
("Zepto","Premium gifting","Bundled ~₹5-6L; Atom ₹30k/mo","10-25%","Low-Med","P3","Expensive entry"),
("Swiggy Instamart","Easiest q-comm test","No fee; weekly POs ₹2-5k","15-25%","Medium","P3","Lightest q-comm commitment"),
]
r=5
for row in Q:
    for ci,v in enumerate(row,1):
        cell=qc.cell(row=r,column=ci,value=v); cell.border=border
        cell.font=Font(name=FONT,size=9,color="222222"); cell.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if ci==6 else "left"))
    pcell=qc.cell(row=r,column=6); pcell.font=Font(name=FONT,size=9,bold=True,color="FFFFFF"); pcell.fill=PatternFill("solid",fgColor=PRI.get(row[5],"999999"))
    if r%2==0:
        for ci in [1,2,3,4,5,7]: qc.cell(row=r,column=ci).fill=PatternFill("solid",fgColor="FAFBFC")
    r+=1
qc.freeze_panes="A5"; widths(qc,[18,24,32,16,20,9,34])

# ---------- BUDGET ----------
bd=sheet("Budget Summary")
banner(bd,"BUDGET SUMMARY — by Domain","Live SUMIF from Master Stack. Ad spend included in monthly. Aggressive scenario.",4)
hdr(bd,4,["Domain","Setup (₹)","Monthly (₹)","3-Month Run (₹)"])
domains=["Brand & Identity","Website & E-commerce","Content & Creative","AI & Creative","Performance Marketing","CRM & Retention","Chat & Social Automation","AI & Automation","Sourcing & Manufacturing","Packaging & Unboxing","Inventory & Warehouse","Order & Logistics","Q-Commerce & Marketplaces","Influencer & PR","Retail / Offline","Finance & Compliance","Legal & IP","HR & Team Ops","Project & Internal Ops","Analytics & BI"]
r=5
for d in domains:
    bd.cell(row=r,column=1,value=d).font=Font(name=FONT,size=9,color="222222")
    bd.cell(row=r,column=2,value=f"=SUMIF('Master Stack'!C:C,A{r},'Master Stack'!J:J)").number_format=RUPEE
    bd.cell(row=r,column=3,value=f"=SUMIF('Master Stack'!C:C,A{r},'Master Stack'!K:K)").number_format=RUPEE
    bd.cell(row=r,column=4,value=f"=B{r}+C{r}*3").number_format=RUPEE
    for ci in range(1,5):
        bd.cell(row=r,column=ci).border=border; bd.cell(row=r,column=ci).alignment=Alignment(horizontal=("left" if ci==1 else "right"),vertical="center")
    if r%2==0:
        for ci in range(1,5): bd.cell(row=r,column=ci).fill=PatternFill("solid",fgColor="FAFBFC")
    r+=1
bd.cell(row=r,column=1,value="TOTAL").font=Font(name=FONT,bold=True,size=10,color="FFFFFF"); bd.cell(row=r,column=1).fill=PatternFill("solid",fgColor=NAVY)
bd.cell(row=r,column=1).alignment=Alignment(horizontal="left",indent=1)
for ci,col in zip((2,3,4),("B","C","D")):
    cell=bd.cell(row=r,column=ci,value=f"=SUM({col}5:{col}{r-1})"); cell.number_format=RUPEE
    cell.font=Font(name=FONT,bold=True,size=10,color=NAVY); cell.fill=PatternFill("solid",fgColor=YELLOW); cell.alignment=Alignment(horizontal="right")
note=bd.cell(row=r+2,column=1,value="In-house work by Arihant/Designomics is shown at internal tool cost, not agency price — that spread is the group's margin advantage. Ad + AI-tool lines are floors; scale with ROAS. AI tools billed in USD are approximated here in INR via their Master Stack lines.")
note.font=Font(name=FONT,italic=True,size=9,color="555555"); note.alignment=Alignment(wrap_text=True,vertical="top")
bd.merge_cells(start_row=r+2,start_column=1,end_row=r+3,end_column=4)
widths(bd,[30,18,18,20])

# order
order=["Overview","Master Stack","Aurra Hype","Designomics","Loop In (rebrand)","Arihant Digital","AI & Creative Stack","Automation Catalogue","Founder OS","Benchmarks","90-Day Roadmap","Q-Commerce & Marketplaces","Budget Summary"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.save(P)
print("all sheets saved:", wb.sheetnames)
