# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CRIMSON="D2042D"; CREAM="EAE5D4"; DARK="2A1A1D"; INK="3A2E30"
CREAM_LT="F4F1E6"; GOLD="F6E7A6"; ROW_ALT="F7F3E8"
FONT="Calibri"
thin=Side(style="thin",color="D8CFC0"); border=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=Workbook()
def sheet(name):
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False; return ws
def banner(ws,title,sub,ncols):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
    c=ws.cell(row=1,column=1,value=title); c.font=Font(name=FONT,bold=True,size=15,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor=CRIMSON); c.alignment=Alignment(horizontal="left",vertical="center",indent=1); ws.row_dimensions[1].height=30
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncols)
    c=ws.cell(row=2,column=1,value=sub); c.font=Font(name=FONT,italic=True,size=9,color=INK)
    c.fill=PatternFill("solid",fgColor=CREAM); c.alignment=Alignment(horizontal="left",indent=1); ws.row_dimensions[2].height=17
def hdr(ws,row,headers):
    for i,h in enumerate(headers,1): ws.cell(row=row,column=i,value=h)
    for cc in range(1,len(headers)+1):
        cell=ws.cell(row=row,column=cc); cell.font=Font(name=FONT,bold=True,color="FFFFFF",size=9)
        cell.fill=PatternFill("solid",fgColor=CRIMSON); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
def widths(ws,ws_w):
    for i,w in enumerate(ws_w,1): ws.column_dimensions[get_column_letter(i)].width=w

# ---- standard category tab: rows = (service, delivers, best_for, engagement, benchmark)
def cat(name, title, sub, rows):
    ws=sheet(name)
    banner(ws,title,sub,7)
    hdr(ws,4,["#","Service / Deliverable","What it delivers","Best for","Engagement","Market Benchmark","Our Price (₹)"])
    aligns=["center","left","left","center","center","left","center"]
    r=5
    for i,row in enumerate(rows,1):
        svc,dl,bf,eng,bm=row
        vals=[i,svc,dl,bf,eng,bm,""]
        for ci,v in enumerate(vals,1):
            cell=ws.cell(row=r,column=ci,value=v); cell.border=border
            cell.font=Font(name=FONT,size=9,color=INK,bold=(ci==2))
            cell.alignment=Alignment(vertical="top",wrap_text=True,horizontal=aligns[ci-1])
        ws.cell(row=r,column=7).fill=PatternFill("solid",fgColor=GOLD)
        if r%2==0:
            for ci in [1,2,3,4,5,6]: ws.cell(row=r,column=ci).fill=PatternFill("solid",fgColor=ROW_ALT)
        r+=1
    ws.freeze_panes="A5"; widths(ws,[4,30,46,13,15,22,15])
    return ws

# ================= OVERVIEW =================
ws=wb.active; ws.title="Ospyr — Start Here"; ws.sheet_view.showGridLines=False
banner(ws,"OSPYR — GROWTH PARTNER SERVICE CATALOGUE","One team to build & scale all four brands: Aurra Hype · Designomics India · Loop In Events · Arihant Digital",2)
rows=[
 ("WHO THIS IS FOR",""),
 ("One founder. Four brands.","You run four businesses alone. Ospyr becomes your single growth partner — we build, run and automate the parts that don't need to be you, so you scale all four without burning out."),
 ("Together or separately","Every service here can be bought per-brand or across the group. Group bundles cost less and share content, data and automations."),
 ("",""),
 ("WHAT'S INSIDE (each tab = a service area)",""),
 ("Branding & Identity","Names, logos, brand systems, trademarks."),
 ("Website & E-Commerce","Stores, landing pages, checkout, conversion."),
 ("Marketing, SEO & GEO","Search, AI-answer visibility, ads, Google presence."),
 ("Social Media & Content","Content engine, reels, community, influencers."),
 ("Automations Menu","Done-for-you automations — described by what they DO for you."),
 ("Founder OS & Dashboards","One command center to run all four brands."),
 ("Sourcing, Manufacturing & Packaging","Supply, production, materials, unboxing."),
 ("Inventory, Delivery & Logistics","Stock, orders, shipping, returns."),
 ("Quick Commerce & Marketplaces","Amazon, Flipkart, quick-commerce, more."),
 ("Finance, Legal & Compliance","Books, GST, contracts, IP."),
 ("HR & Operations","Hiring, payroll, SOPs, team systems."),
 ("International Expansion","Take the brands global."),
 ("Bundles & Engagement","Suggested packages + how we work together."),
 ("",""),
 ("HOW TO READ THE PRICING",""),
 ("Market Benchmark","What the market typically charges for this service in India (2026) — your reference point."),
 ("Our Price (₹)","Left blank on purpose. This is where Ospyr's quote goes — filled in per-brand or as a group bundle."),
 ("Note","Ad spend, platform subscriptions and third-party fees are separate from service fees. Prices are indicative and finalised per scope."),
]
r=4
for a,b in rows:
    ca=ws.cell(row=r,column=1,value=a); cb=ws.cell(row=r,column=2,value=b)
    if b=="" and a:
        ca.font=Font(name=FONT,bold=True,size=12,color=CRIMSON); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
        ca.fill=PatternFill("solid",fgColor=CREAM)
    else:
        ca.font=Font(name=FONT,bold=True,size=10,color=CRIMSON); cb.font=Font(name=FONT,size=10,color=INK)
    ca.alignment=Alignment(vertical="top",wrap_text=True); cb.alignment=Alignment(vertical="top",wrap_text=True)
    r+=1
widths(ws,[30,100])

# ================= BRANDING =================
cat("Branding & Identity","BRANDING & IDENTITY","Names, logos, brand systems and legal ownership across all four brands.",[
 ("Brand naming & architecture","New/refined names + how your four brands relate as a group","All 4","One-time","₹40,000–1,50,000"),
 ("Logo & visual identity","Logo suite, colours, typography, brand guidelines","Per brand","One-time","₹25,000–1,50,000"),
 ("Rebrand (Loop In)","Full rename + identity to fix the existing name clash","Loop In","One-time","₹60,000–2,00,000"),
 ("Brand voice & messaging","Tone, taglines, key messages used everywhere","Per brand","One-time","₹15,000–60,000"),
 ("Trademark search & filing","Own your names/logos legally; block copycats","All 4","Per mark","₹8,000–15,000 + govt fee"),
 ("Brand collateral kit","Business cards, letterheads, signatures, templates","Per brand","One-time","₹10,000–40,000"),
 ("Product / packaging brand art","On-brand artwork for boxes, labels, tags","Aurra/Design","Per project","₹15,000–75,000"),
 ("Brand refresh (annual)","Keep identity current as you scale","Per brand","Yearly","₹20,000–1,00,000"),
])

# ================= WEB =================
cat("Website & E-Commerce","WEBSITE & E-COMMERCE","High-converting stores, sites and checkouts built and optimised for sales.",[
 ("E-commerce store build","Fast, mobile-first online store designed to convert","Aurra/Design","One-time","₹40,000–3,00,000"),
 ("Store replatform / upgrade","Move off a limiting platform; faster, more capable store","Designomics","One-time","₹40,000–1,50,000"),
 ("Custom gift-box / product builder","Let customers build their own box/product online","Designomics","One-time","₹40,000–1,50,000"),
 ("Business / lead-gen website","Portfolio, packages and enquiry capture site","Loop In/Arihant","One-time","₹25,000–1,00,000"),
 ("Landing pages & funnels","Fast campaign/drop pages that convert paid traffic","All 4","Per page","₹8,000–30,000"),
 ("Conversion optimisation (CRO)","Improve add-to-cart & checkout; more sales, same traffic","Aurra/Design","Monthly","₹20,000–75,000/mo"),
 ("One-tap / UPI fast checkout","Reduce drop-off with instant UPI/COD checkout","Aurra/Design","Setup + Monthly","₹15,000 + usage"),
 ("Reviews, upsell & personalisation","Higher order value + trust on every visit","Aurra/Design","Setup + Monthly","₹15,000 + ₹4,000/mo"),
 ("Website care & hosting","Speed, security, updates, uptime","All 4","Monthly","₹5,000–25,000/mo"),
 ("Google Workspace & domain setup","Professional email, domains, DNS, accounts","All 4","One-time","₹5,000–15,000 + subs"),
])

# ================= MARKETING SEO GEO =================
cat("Marketing, SEO & GEO","MARKETING, SEO & GEO","Get found on Google, in AI answers, and through paid ads that pay back.",[
 ("Search engine optimisation (SEO)","Rank on Google for the terms buyers search","All 4","Monthly","₹25,000–1,50,000/mo"),
 ("GEO / AI-answer optimisation","Get your brand cited by AI assistants & AI search","All 4","Monthly","₹15,000–60,000/mo"),
 ("Google Search Console & presence","Set up + manage Google search health & local presence","All 4","Setup + Monthly","₹8,000 + ₹5,000/mo"),
 ("Google Business Profile & local SEO","Store footfall + local event/gifting leads","Aurra/Loop In","Monthly","₹5,000–20,000/mo"),
 ("Performance ads management","Run & scale paid ads profitably (fee only; spend separate)","All 4","Monthly","15–20% of spend, min ₹30,000/mo"),
 ("Shopping / catalog ads","Product ads that show your items to ready buyers","Aurra/Design","Monthly","₹25,000–75,000/mo"),
 ("Marketplace ads","Sponsored placements on Amazon/Flipkart etc.","Aurra/Design","Monthly","₹20,000–60,000/mo"),
 ("Creative testing & analytics","Find winning ads faster; lower cost per sale","All 4","Monthly","₹20,000–60,000/mo"),
 ("Profit & attribution reporting","See true profit per ad/channel after all costs","Aurra/Design","Monthly","₹15,000–40,000/mo"),
 ("Full-funnel growth retainer","End-to-end marketing across all channels","Per brand","Monthly","₹75,000–5,00,000/mo"),
])

# ================= SOCIAL =================
cat("Social Media & Content","SOCIAL MEDIA & CONTENT","A content engine that keeps every brand consistently visible and selling.",[
 ("Social media management","Strategy, calendar, posting across all platforms","Per brand","Monthly","₹20,000–80,000/mo"),
 ("Short-form video / reels engine","3–5 scroll-stopping reels per week per brand","Per brand","Monthly","₹25,000–1,00,000/mo"),
 ("Product & lifestyle photography","Sales-grade catalog + campaign visuals","Aurra/Design","Per shoot","₹20,000–1,00,000"),
 ("Content repurposing","Turn one shoot into 20+ posts/clips","All 4","Monthly","₹15,000–50,000/mo"),
 ("Community & DM management","Fast replies that turn followers into buyers","All 4","Monthly","₹15,000–50,000/mo"),
 ("Influencer & creator campaigns","Seeding + paid collabs with the right creators","Aurra Hype","Per campaign","₹50,000–5,00,000 + fees"),
 ("User-generated content (UGC) pipeline","Steady stream of authentic creator content","Aurra/Design","Monthly","₹30,000–1,00,000/mo"),
 ("Drop / launch campaigns","Hype, waitlists and buzz around each product drop","Aurra Hype","Per drop","₹25,000–1,00,000"),
 ("PR & media outreach","Press features for trust + backlinks","All 4","Monthly","₹20,000–75,000/mo"),
])

# ================= AUTOMATIONS (special) =================
au=sheet("Automations Menu")
banner(au,"AUTOMATIONS MENU — done-for-you, described by what they DO","You don't manage tools. You click, and it happens. Set-up once, runs forever. (Sample menu — ask for more.)",7)
hdr(au,4,["#","Automation","What happens (the outcome for you)","Triggered by","Best for","Market Benchmark","Our Price (₹)"])
AU=[
("One-click visual → posted","You pick or click an image; it's auto-polished into a branded visual and posted across your social channels — no editor, no scheduling","You / a click","All 4","₹15,000 setup + ₹6,000/mo"),
("New product → live everywhere","Add a product once; visuals, description and ads are drafted and it's listed across your store and marketplaces automatically","New product added","Aurra/Design","₹25,000 setup + ₹8,000/mo"),
("Abandoned cart → recovered","A shopper leaves without buying; they get a friendly WhatsApp + email with the product and an offer, and many come back","Cart left behind","Aurra/Design","₹12,000 setup + ₹5,000/mo"),
("Order updates on WhatsApp","Every buyer is auto-updated on WhatsApp at each step — confirmed, shipped, out for delivery, delivered — cutting 'where's my order' messages","Order status change","All 4","₹10,000 setup + ₹4,000/mo"),
("Comment → auto-DM sale","Someone comments a keyword on your post; they instantly get a DM with the link/offer and are captured as a lead","Comment on a post","Aurra/Design","₹12,000 setup + ₹5,000/mo"),
("Instagram DMs answered 24x7","Every DM gets an instant on-brand reply — sizing, price, availability — and hot buyers are flagged to you","Instagram DM","All 4","₹15,000 setup + ₹6,000/mo"),
("WhatsApp shop & checkout","Customers browse your catalog, order and pay entirely inside WhatsApp","Customer message","Aurra/Design","₹18,000 setup + ₹6,000/mo"),
("Review & photo request","A few days after delivery, buyers are asked for a review + photo and rewarded — building trust that lowers ad costs","Order delivered","Aurra/Design","₹10,000 setup + ₹4,000/mo"),
("Occasion gifting reminders","Customers are nudged before birthdays/festivals with a ready-to-buy gift — turning one-time buyers into repeat revenue","Date approaching","Designomics","₹12,000 setup + ₹5,000/mo"),
("Lead → instant follow-up","Any enquiry (form, DM, ad) is logged, answered instantly with a booking link, and assigned — no lead ever dropped","New lead","Loop In/Design","₹15,000 setup + ₹6,000/mo"),
("Telegram VIP drop alerts","Your VIP list gets first access to every drop via an automated broadcast","New drop","Aurra Hype","₹8,000 setup + ₹3,000/mo"),
("Daily business digest","Every morning you get one message: yesterday's sales, ad results, best sellers and low stock — across all four brands","Every morning","All 4","₹20,000 setup + ₹5,000/mo"),
("Low-stock → auto reorder","When a winner runs low, you're alerted and a purchase order is drafted for one-tap approval — no stockouts","Stock below threshold","All 4","₹15,000 setup + ₹4,000/mo"),
("Risky COD → auto-verify","High-risk cash-on-delivery orders are automatically verified before shipping — protecting your margin from returns","COD order placed","Aurra/Design","₹12,000 setup + ₹5,000/mo"),
("Post-event feedback & upsell","After every event, guests get a feedback + testimonial request and your next offer — reviews and rebookings on autopilot","Event ends","Loop In","₹10,000 setup + ₹4,000/mo"),
("Auto social scheduling","Any new content is automatically formatted and scheduled across all your channels at the best times","New content ready","All 4","₹10,000 setup + ₹4,000/mo"),
("Meeting → action list","Every call is auto-summarised into notes and tasks so nothing slips","Call ends","All 4","₹8,000 setup + ₹3,000/mo"),
("Weekly performance report","Every Monday, a clean branded report of each brand's numbers lands in your inbox — no spreadsheets","Every Monday","All 4","₹12,000 setup + ₹4,000/mo"),
("Influencer outreach on autopilot","The right creators are found and contacted for you every week, with replies tracked","Weekly","Aurra Hype","₹15,000 setup + ₹6,000/mo"),
("Invoice & payment chasing","Invoices are sent, tracked and politely chased until paid — for B2B and corporate deals","Invoice due","Loop In/Design","₹10,000 setup + ₹4,000/mo"),
("Custom automation (built to order)","Describe any repetitive task; we build an automation that does it for you","On request","All 4","From ₹15,000 / automation"),
]
aligns=["center","left","left","center","center","left","center"]
r=5
for i,row in enumerate(AU,1):
    svc,dl,trg,bf,bm=row
    vals=[i,svc,dl,trg,bf,bm,""]
    for ci,v in enumerate(vals,1):
        cell=au.cell(row=r,column=ci,value=v); cell.border=border
        cell.font=Font(name=FONT,size=9,color=INK,bold=(ci==2)); cell.alignment=Alignment(vertical="top",wrap_text=True,horizontal=aligns[ci-1])
    au.cell(row=r,column=7).fill=PatternFill("solid",fgColor=GOLD)
    if r%2==0:
        for ci in [1,2,3,4,5,6]: au.cell(row=r,column=ci).fill=PatternFill("solid",fgColor=ROW_ALT)
    r+=1
au.freeze_panes="A5"; widths(au,[4,26,50,16,12,22,15])

# ================= FOUNDER OS =================
cat("Founder OS & Dashboards","FOUNDER OS & DASHBOARDS","One command center to run all four brands — most of it on autopilot.",[
 ("Founder command dashboard","One live screen: sales, ad results, cash, stock, leads — all four brands","All 4","Setup + Monthly","₹40,000 setup + ₹10,000/mo"),
 ("Daily founder digest","Your business in one morning message (see Automations)","All 4","Monthly","₹5,000/mo"),
 ("Ask-your-business assistant","Ask questions in plain language and get answers from your own live data","All 4","Setup + Monthly","₹25,000 setup + ₹8,000/mo"),
 ("Unified work & task system","Every brand's projects, tasks and SOPs in one organised place","All 4","Setup + Monthly","₹20,000 setup + ₹5,000/mo"),
 ("Lead & sales pipeline (CRM)","Track every enquiry from first contact to paid","Loop In/Design","Setup + Monthly","₹20,000 setup + ₹6,000/mo"),
 ("Invoicing & payments tracking","Create, send and track invoices; see who owes what","Loop In/Design","Monthly","₹5,000–15,000/mo"),
 ("Finance snapshot & cash view","Live view of income, spend, burn and runway","All 4","Monthly","₹10,000–25,000/mo"),
 ("Time & focus system","Your week auto-planned around priorities and deep work","Founder","Monthly","₹5,000/mo"),
 ("One-tap approvals","Creative, spend and hires come to you as approve/reject","All 4","Setup + Monthly","included in dashboard"),
 ("Weekly scorecard","Targets vs actuals per brand, built and sent automatically","All 4","Monthly","₹5,000/mo"),
])

# ================= SOURCING / MANUFACTURING / PACKAGING =================
cat("Sourcing, Mfg & Packaging","SOURCING, MANUFACTURING & PACKAGING","Get products made and delivered beautifully — supply, production and unboxing.",[
 ("Supplier & manufacturer sourcing","Find, vet and negotiate reliable production partners","Aurra/Design","Per project","₹15,000–75,000"),
 ("Apparel production management","Manage cut-and-sew, blanks and small-batch runs for drops","Aurra Hype","Monthly","₹25,000–75,000/mo"),
 ("On-demand printing setup","Print products only when ordered — no big minimums","Aurra/Design","Setup","₹10,000–40,000"),
 ("Merch & gifting supply network","Reliable supply for mugs, bottles, candles, kits, etc.","Designomics","Per project","₹15,000–60,000"),
 ("Quality control system","Checklists + tech-packs for consistent quality, fewer returns","Aurra/Design","Setup + Monthly","₹15,000 + ₹8,000/mo"),
 ("Custom packaging design","Branded boxes, mailers, tape and inserts","All 4","Per project","₹20,000–1,00,000"),
 ("Unboxing experience","Inserts, thank-you cards, stickers and QR that drive reviews & referrals","Aurra/Design","Setup","₹15,000–50,000"),
 ("Premium / sustainable packaging","Eco + premium line that raises perceived value","All 4","Per project","₹25,000–1,25,000"),
 ("Material & inventory sourcing","Source raw materials and stock at the right price","All 4","Per project","₹10,000–50,000"),
])

# ================= INVENTORY / DELIVERY / LOGISTICS =================
cat("Inventory & Logistics","INVENTORY, DELIVERY & LOGISTICS","Stock, orders, shipping and returns — organised and mostly automated.",[
 ("Inventory management system","One live stock count across store, marketplaces and shop","All 4","Setup + Monthly","₹15,000 setup + ₹5,000/mo"),
 ("Barcode / SKU & labelling","Scan-based picking; stops overselling","All 4","Setup","₹10,000–40,000"),
 ("Warehouse / stockroom setup","Organised bins and flow for fast fulfilment","All 4","One-time","₹40,000–1,50,000"),
 ("Multi-channel stock sync","Stock updates everywhere automatically when something sells","Aurra/Design","Monthly","₹8,000–20,000/mo"),
 ("Shipping & courier setup","Best courier per pincode; tracking; failed-delivery handling","All 4","Setup + Monthly","₹8,000 + ₹4,000/mo"),
 ("Returns / exchange system","Smooth returns that protect margin and customer trust","Aurra/Design","Setup + Monthly","₹10,000 + ₹3,000/mo"),
 ("Cash-on-delivery / RTO control","Cut losses from returned/undelivered COD orders","Aurra/Design","Monthly","₹10,000–25,000/mo"),
 ("Order management system","All orders from every channel in one queue","All 4","Setup + Monthly","₹15,000 + ₹6,000/mo"),
 ("Same-day / hyperlocal delivery","Local same-day delivery for store & city orders","Aurra Hype","Monthly","₹8,000–20,000/mo"),
])

# ================= QUICK COMMERCE / MARKETPLACES =================
cat("Quick Commerce & Marketplaces","QUICK COMMERCE & MARKETPLACES","Sell everywhere your customers already shop.",[
 ("Amazon store setup & management","Listings, enhanced content, ads and account health","Aurra/Design","Setup + Monthly","₹25,000 setup + ₹20,000/mo"),
 ("Flipkart store setup & management","Full onboarding, catalog and growth","Aurra/Design","Setup + Monthly","₹20,000 setup + ₹18,000/mo"),
 ("Fashion marketplace onboarding","Myntra / Ajio listing and management for apparel","Aurra Hype","Setup + Monthly","₹25,000 setup + ₹20,000/mo"),
 ("Value-marketplace listing","High-volume, low-price gifting/merch channel","Designomics","Setup + Monthly","₹15,000 setup + ₹10,000/mo"),
 ("Quick-commerce onboarding","List fast-moving SKUs on quick-delivery apps","Designomics","Project + Monthly","₹40,000 setup + platform fees"),
 ("Global handmade marketplace","Sell stickers/gifting internationally","Designomics","Setup + Monthly","₹15,000 setup + ₹8,000/mo"),
 ("Marketplace catalog & content","Photos, A+ content and copy that win the buy box","Aurra/Design","Per catalog","₹15,000–60,000"),
 ("Channel strategy & mix","Which channels to be on and in what order","All 4","One-time","₹20,000–50,000"),
])

# ================= FINANCE / LEGAL =================
cat("Finance, Legal & Compliance","FINANCE, LEGAL & COMPLIANCE","Clean books, protected brand, and full compliance across all entities.",[
 ("Accounting & bookkeeping","Clean, investor-ready books across all brands","All 4","Monthly","₹8,000–30,000/mo"),
 ("GST registration & filing","Registration + monthly/quarterly returns","All 4","Monthly","₹3,000–15,000/mo"),
 ("Business setup & structuring","Company/LLP setup, structure across brands, export code","All 4","One-time","₹15,000–75,000"),
 ("Payment gateway & reconciliation","Accept UPI/cards/COD; auto-match payouts","All 4","Setup","₹5,000–20,000"),
 ("Unit economics & P&L dashboard","Know your true profit, CAC and margins per brand","All 4","Setup + Monthly","₹15,000 + ₹5,000/mo"),
 ("Contracts & policies","Supplier, influencer, employment and store policies","All 4","Per pack","₹10,000–50,000"),
 ("Trademark & IP protection","Register and defend your brand assets","All 4","Per mark","₹8,000–15,000 + fees"),
 ("Data-privacy compliance","Meet India (and global) privacy rules","All 4","Setup","₹10,000–40,000"),
 ("Expense & spend control","Track and control burn across four brands","All 4","Monthly","₹5,000–15,000/mo"),
])

# ================= HR / OPS =================
cat("HR & Operations","HR & OPERATIONS","Build the team and the systems that let the group run without you.",[
 ("Hiring & recruitment","Find and onboard marketing, ops and design talent","All 4","Per hire","₹15,000–75,000 / hire"),
 ("Payroll & attendance","Compliant salaries, attendance and statutory filings","All 4","Monthly","₹3,000–12,000/mo"),
 ("SOPs & training library","Documented processes so anyone can run the playbook","All 4","One-time","₹20,000–1,00,000"),
 ("Freelancer / contractor management","Manage and pay designers, editors and creators","All 4","Monthly","₹5,000–15,000/mo"),
 ("Ops audit & systemisation","Find bottlenecks; automate or delegate them","All 4","Per project","₹25,000–1,00,000"),
 ("Team dashboards & accountability","Everyone sees their targets and progress","All 4","Setup + Monthly","₹15,000 + ₹5,000/mo"),
])

# ================= INTERNATIONAL =================
cat("International Expansion","INTERNATIONAL EXPANSION","Take the brands global — the right markets, channels and setup.",[
 ("International market research","Which markets fit each brand and why","Per brand","One-time","₹25,000–1,00,000"),
 ("Cross-border store & payments","Sell and get paid in other countries","Aurra/Design","Setup","₹30,000–1,50,000"),
 ("Global marketplace onboarding","List on international marketplaces","Aurra/Design","Setup + Monthly","₹25,000 + ₹15,000/mo"),
 ("International shipping & fulfilment","Reliable, cost-effective global delivery","Aurra/Design","Setup + Monthly","₹15,000 + usage"),
 ("Multi-currency & tax setup","Handle currencies, duties and overseas tax","All 4","Setup","₹20,000–75,000"),
 ("Localised content & ads","Content and campaigns tuned to each market","Per brand","Monthly","₹40,000–1,50,000/mo"),
 ("Global brand & trademark","Protect the brand in target countries","All 4","Per market","₹25,000–1,00,000 + fees"),
])

# ================= BUNDLES =================
bd=sheet("Bundles & Engagement")
banner(bd,"BUNDLES & ENGAGEMENT — suggested ways to work together","Mix per-brand or group-wide. Group bundles share content, data and automations = more value, lower cost.",4)
hdr(bd,4,["Bundle / Engagement","What's included","Best for","Our Price (₹) — you fill"])
BD=[
 ("Launch Pack (per brand)","Brand + store/site + core automations + social setup — everything to go live","A single brand starting up","", ),
 ("Growth Engine (per brand)","Ads + SEO/GEO + content + retention + creative, managed monthly","A brand ready to scale","", ),
 ("Automation Suite","10–15 done-for-you automations + Founder dashboard","Any brand drowning in manual work","", ),
 ("Founder OS (group)","One command center + daily digest + assistant across all four brands","The founder running everything","", ),
 ("Full-Service Group Retainer","End-to-end across all four brands — build, market, automate, operate","One partner for everything","", ),
 ("Store & Marketplace Pack","D2C store + Amazon/Flipkart + quick-commerce onboarding","Products to sell everywhere","", ),
 ("Content & Social Pack","Reels engine + photography + social management + UGC","Brands needing constant content","", ),
 ("Operations Pack","Sourcing + packaging + inventory + logistics + returns","Scaling physical fulfilment","", ),
 ("International Pack","Market research + cross-border store + global channels","Ready to go worldwide","", ),
 ("Custom Scope","Pick any services from any tab; we quote as one plan","Anything not covered above","", ),
]
r=5
for row in BD:
    for ci,v in enumerate(row,1):
        cell=bd.cell(row=r,column=ci,value=v); cell.border=border
        cell.font=Font(name=FONT,size=10,color=INK,bold=(ci==1)); cell.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("left" if ci in(1,2,3) else "center"))
    bd.cell(row=r,column=4).fill=PatternFill("solid",fgColor=GOLD)
    if r%2==0:
        for ci in [1,2,3]: bd.cell(row=r,column=ci).fill=PatternFill("solid",fgColor=ROW_ALT)
    r+=1
# how we work
r+=1
bd.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
c=bd.cell(row=r,column=1,value="HOW WE WORK"); c.font=Font(name=FONT,bold=True,size=12,color=CRIMSON); c.fill=PatternFill("solid",fgColor=CREAM); r+=1
for step in ["1 · Discovery — we map your four brands, goals and gaps.","2 · Proposal — we pick services + fill in Our Price as a per-brand or group plan.","3 · Build — we set up brand, store, automations and systems.","4 · Run & scale — we manage growth and report weekly.","5 · Review — monthly numbers, next priorities, new automations."]:
    cc=bd.cell(row=r,column=1,value=step); cc.font=Font(name=FONT,size=10,color=INK); bd.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
    cc.alignment=Alignment(wrap_text=True,vertical="top"); r+=1
widths(bd,[28,52,22,22])

# order
order=["Ospyr — Start Here","Branding & Identity","Website & E-Commerce","Marketing, SEO & GEO","Social Media & Content","Automations Menu","Founder OS & Dashboards","Sourcing, Mfg & Packaging","Inventory & Logistics","Quick Commerce & Marketplaces","Finance, Legal & Compliance","HR & Operations","International Expansion","Bundles & Engagement"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.save("/sessions/bold-epic-ramanujan/mnt/outputs/Ospyr_Service_Catalogue_Pitch.xlsx")
print("saved. sheets:", len(wb.sheetnames))
