# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F2A44"; BLUE="2E5AAC"; TEAL="0E7C7B"; ORANGE="C05621"; PURPLE="5B3A8C"
LT_BLUE="E8EEF9"; LT_GREY="F4F5F7"; YELLOW="FFF3C4"; GREEN="1E7D34"; RED="B02A37"
FONT="Calibri"; RUPEE='"₹"#,##0'
thin=Side(style="thin",color="C9CED6"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
P="/sessions/bold-epic-ramanujan/mnt/outputs/Harshdeep_Group_Stack.xlsx"
wb=load_workbook(P)

# fix typo in Master Stack vendor cell
ms=wb["Master Stack"]
for row in ms.iter_rows(min_col=9,max_col=9):
    cell=row[0]
    if isinstance(cell.value,str) and "Bizongo/Pack" in cell.value:
        cell.value="Designomics + Bizongo / Packman"

def head(ws,text,sub,ncols):
    ws.sheet_view.showGridLines=False
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
    c=ws.cell(row=1,column=1,value=text); c.font=Font(name=FONT,bold=True,size=15,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[1].height=28
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncols)
    c=ws.cell(row=2,column=1,value=sub); c.font=Font(name=FONT,italic=True,size=9,color="333333")
    c.fill=PatternFill("solid",fgColor=LT_GREY); c.alignment=Alignment(horizontal="left",indent=1); ws.row_dimensions[2].height=16
def hdr(ws,row,headers,color=NAVY):
    for i,h in enumerate(headers,1): ws.cell(row=row,column=i,value=h)
    for c in range(1,len(headers)+1):
        cell=ws.cell(row=row,column=c); cell.font=Font(name=FONT,bold=True,color="FFFFFF",size=10)
        cell.fill=PatternFill("solid",fgColor=color); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border

# ================= AI AUTOMATIONS =================
ai=wb.create_sheet("AI Automations")
head(ai,"AI & AUTOMATION CATALOGUE","Build on n8n/Make + WhatsApp API + GPT/Claude. Each = trigger → action → impact.",7)
H=["#","Automation","Business","Trigger","What it does","Tools","Impact"]
hdr(ai,4,H)
A=[
("Abandoned-cart WhatsApp+email recovery","Aurra Hype / Designomics","Checkout started, not completed","Sends timed WhatsApp + email with product image & discount","Shopify + Klaviyo + AiSensy + n8n","Recovers 8-15% lost carts"),
("Order status auto-updates on WhatsApp","All","Order placed / shipped / OFD / delivered","Auto WhatsApp with tracking at each stage","Shiprocket + WhatsApp API + n8n","Cuts 'where is my order' tickets ~60%"),
("AI support chatbot (site + WhatsApp)","All","Customer message","Answers sizing, returns, order status; hands off to human","Chatbase/Intercom Fin + order API","24x7 support, faster conversion"),
("Review + UGC request flow","Aurra Hype / Designomics","Order delivered +5 days","Requests review + photo; rewards with coupon","Judge.me/Loox + WhatsApp + n8n","More social proof, repeat buyers"),
("Daily ops digest to founder","Group","Every morning 8am","Sales, ad spend, ROAS, top SKUs, low stock → WhatsApp","Sheets + GA4 + Meta API + GPT + n8n","One-glance control of 3 brands"),
("Low-stock + reorder alerts","All","Stock < threshold","Alerts + drafts PO to supplier","Zoho Inventory + n8n + WhatsApp","Prevents stockouts on winners"),
("Lead capture → CRM → auto-reply","Loop In / Designomics B2B","Form / DM / ad lead","Logs to CRM, sends intro + booking link, notifies sales","Forms + Zoho CRM + n8n","No lead dropped; faster follow-up"),
("AI creative & caption generator","Group","New product / drop","Generates ad copy, captions, descriptions, hashtags","GPT/Claude + Canva API","Cuts content time ~70%"),
("AI product photo / on-model images","Aurra Hype / Designomics","New SKU added","Generates lifestyle/on-model shots from flat-lay","Photoroom/Flair + Midjourney","Catalog-ready visuals, less shoot cost"),
("RTO / COD risk flagging","All","COD order placed","Scores risk; auto-verify high-risk via WhatsApp","GoKwik / Shiprocket + n8n","Reduces RTO losses"),
("Influencer / UGC outreach at scale","Aurra Hype","Weekly","Finds + DMs relevant micro-creators, tracks replies","Sheets + IG + GPT + n8n","Scales seeding cheaply"),
("Auto marketplace listing sync","Aurra Hype","New/edited product","Pushes SKU to Amazon/Flipkart with A+ copy","Unicommerce + GPT + n8n","List everywhere in minutes"),
("Weekly performance report (client-ready)","Group / Arihant clients","Every Monday","Auto-builds branded PDF report of KPIs","Looker + GPT + n8n","Also a sellable Arihant service"),
("Post-event feedback + upsell (Loop In)","Loop In","Event ends","Sends feedback form + testimonial + next-service offer","Forms + CRM + WhatsApp","Reviews + repeat bookings"),
("Birthday / occasion gifting reminders","Designomics","Date in CRM approaches","Nudges customers to gift; pre-filled cart","CRM + Klaviyo + WhatsApp","Recurring gifting revenue"),
]
r=5
for i,(nm,biz,trg,wht,tl,imp) in enumerate(A,1):
    for ci,v in enumerate([i,nm,biz,trg,wht,tl,imp],1):
        cell=ai.cell(row=r,column=ci,value=v); cell.border=border
        cell.font=Font(name=FONT,size=9,color="222222")
        cell.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if ci==1 else "left"))
    if r%2==0:
        for ci in range(1,8): ai.cell(row=r,column=ci).fill=PatternFill("solid",fgColor="FAFBFC")
    r+=1
ai.freeze_panes="A5"
for i,w in enumerate([4,30,20,24,40,28,30],1): ai.column_dimensions[get_column_letter(i)].width=w

# ================= 90-DAY ROADMAP =================
rd=wb.create_sheet("90-Day Roadmap")
head(rd,"90-DAY SCALE ROADMAP","Month 1 Foundation · Month 2 Engine · Month 3 Scale. 'Proof' = the milestone that shows it's working.",6)
H=["Phase","Business","Workstream","Key Actions","Owner","Proof / KPI by end of phase"]
hdr(rd,4,H)
R=[
("M1 · Foundation (Wk 1-4)","Group","Legal + Brand","Register/clean entities & GST; trademark search+file all 3; rebrand Loop In; unify brand systems","Founder + CA + Designomics","3 TMs filed · Loop In new name live · brand kits done"),
("M1 · Foundation","Group","Tech backbone","Shopify + domains + WhatsApp API + payments + analytics + n8n + inventory + accounting live","Arihant","All core systems live & connected"),
("M1 · Foundation","Aurra Hype","Store + Launch","Upgrade store, shoot catalog, POS live, amplify 12 Jul store launch as content+PR","Arihant + Designomics","Store launch = 500+ footfall / big content spike"),
("M1 · Foundation","Group","Content engine","Photo/video pod running; 3-5 reels/wk per brand; social calendars live","Arihant content pod","90+ assets shipped · consistent posting"),
("M1 · Foundation","Group","Packaging","Branded boxes/mailers + inserts designed & ordered; gift-box configurator live","Designomics","Branded unboxing shipping on every order"),
("M1 · Foundation","Aurra Hype","Paid ads on","Meta always-on live; pixel/CAPI; first creative tests","Arihant","First profitable campaigns · ROAS baseline set"),
("M2 · Engine (Wk 5-8)","Aurra Hype","Scale paid + CRO","Add Google/PMax; scale winners; CRO; email/SMS/WhatsApp flows live","Arihant","ROAS ≥ 2.5 · repeat-rate flows recovering revenue"),
("M2 · Engine","Group","Automations","Deploy cart recovery, order updates, chatbot, ops digest, review flow","Arihant","10+ automations live · support tickets down"),
("M2 · Engine","Aurra Hype","Marketplaces","Launch Amazon + Flipkart with A+ content + ads","Arihant","Live on 2 marketplaces · first marketplace sales"),
("M2 · Engine","Designomics","Corporate gifting","Outbound to corporates; catalog + quote system; festive push","Founder + Arihant","5+ corporate gifting deals in pipeline"),
("M2 · Engine","Loop In","Pipeline","New brand launch; lead-gen ads; CRM; 2-3 signature events booked","Loop In + Arihant","3 events booked · case studies started"),
("M2 · Engine","Group","Retention + loyalty","Loyalty/referral live; win-back flows; UGC pipeline","Arihant","Repeat-purchase rate rising · referral orders"),
("M3 · Scale (Wk 9-12)","Aurra Hype","Scale + drops","Big drop calendar; influencer seeding; scale ad budgets on winners","Arihant + creators","Record revenue month · hero SKUs identified"),
("M3 · Scale","Designomics","Q-commerce + POD","Test Blinkit/Zepto gifting SKUs; POD long-tail; festive campaigns","Founder + Arihant","Q-comm pilot live · gifting revenue scaling"),
("M3 · Scale","Group","Offline + PR","Pop-ups/exhibitions/college activations; PR features","Loop In + Designomics","3 activations · press mentions · email list growth"),
("M3 · Scale","Group","Systemise + productise","SOPs, hires, dashboards; package Arihant offer to sell externally","Founder + Arihant","Team hired · Arihant lands external clients (proof)"),
("M3 · Scale","Group","Prove the numbers","Full unit-economics review; investor/partner deck from real data","Arihant + CA","Documented growth curve · CAC/AOV/ROAS/repeat proof"),
]
r=5; phase_c={"M1":"E8EEF9","M2":"E2EFDA","M3":"FCE4D6"}
for row in R:
    for ci,v in enumerate(row,1):
        cell=rd.cell(row=r,column=ci,value=v); cell.border=border
        cell.font=Font(name=FONT,size=9,color="222222",bold=(ci==1))
        cell.alignment=Alignment(vertical="top",wrap_text=True,horizontal="left")
    key=row[0][:2]
    rd.cell(row=r,column=1).fill=PatternFill("solid",fgColor=phase_c.get(key,"FFFFFF"))
    r+=1
rd.freeze_panes="A5"
for i,w in enumerate([20,13,18,46,20,38],1): rd.column_dimensions[get_column_letter(i)].width=w

# ================= BUDGET SUMMARY =================
bd=wb.create_sheet("Budget Summary")
head(bd,"BUDGET SUMMARY — by Domain","Rolled up live from Master Stack (SUMIF). Ad spend included in monthly. Aggressive-spend scenario.",4)
H=["Domain","Setup (₹, one-time)","Monthly (₹)","3-Month Run (₹)"]
hdr(bd,4,H)
domains=["Brand & Identity","Website & E-commerce","Content & Creative","Performance Marketing","Organic & Social","SEO / GEO / AEO","CRM & Retention","AI & Automation","Customer Support","Sourcing & Manufacturing","Packaging & Unboxing","Inventory & Warehouse","Order & Logistics","Q-Commerce & Marketplaces","Retail / Offline","Finance & Compliance","Legal & IP","HR & Team Ops","Project & Internal Ops","Analytics & BI"]
r=5
for d in domains:
    bd.cell(row=r,column=1,value=d).font=Font(name=FONT,size=9,color="222222")
    bd.cell(row=r,column=2,value=f"=SUMIF('Master Stack'!C:C,A{r},'Master Stack'!J:J)").number_format=RUPEE
    bd.cell(row=r,column=3,value=f"=SUMIF('Master Stack'!C:C,A{r},'Master Stack'!K:K)").number_format=RUPEE
    bd.cell(row=r,column=4,value=f"=B{r}+C{r}*3").number_format=RUPEE
    for ci in range(1,5):
        bd.cell(row=r,column=ci).border=border
        bd.cell(row=r,column=ci).alignment=Alignment(horizontal=("left" if ci==1 else "right"),vertical="center")
    if r%2==0:
        for ci in range(1,5): bd.cell(row=r,column=ci).fill=PatternFill("solid",fgColor="FAFBFC")
    r+=1
# total
bd.cell(row=r,column=1,value="TOTAL").font=Font(name=FONT,bold=True,size=10,color="FFFFFF")
bd.cell(row=r,column=1).fill=PatternFill("solid",fgColor=NAVY)
for ci,col in zip((2,3,4),("B","C","D")):
    cell=bd.cell(row=r,column=ci,value=f"=SUM({col}5:{col}{r-1})"); cell.number_format=RUPEE
    cell.font=Font(name=FONT,bold=True,size=10,color=NAVY); cell.fill=PatternFill("solid",fgColor=YELLOW)
    cell.alignment=Alignment(horizontal="right")
bd.cell(row=r,column=1).alignment=Alignment(horizontal="left",indent=1)
note=bd.cell(row=r+2,column=1,value="Note: 'In-house' work by Arihant/Designomics is shown at internal tool cost, not agency price — that spread is the group's structural margin advantage. Ad budgets scale with ROAS; treat monthly ad lines as starting floors, not caps.")
note.font=Font(name=FONT,italic=True,size=9,color="555555"); note.alignment=Alignment(wrap_text=True,vertical="top")
bd.merge_cells(start_row=r+2,start_column=1,end_row=r+3,end_column=4)
for i,w in enumerate([28,20,18,20],1): bd.column_dimensions[get_column_letter(i)].width=w

# ================= Q-COMMERCE & MARKETPLACES =================
qc=wb.create_sheet("Q-Commerce & Marketplaces")
head(qc,"Q-COMMERCE & MARKETPLACE CHANNELS","Real 2026 India onboarding costs & fit. Verify current terms at signup — platforms change fast.",7)
H=["Channel","Best for","Onboarding cost / model","Commission","Fit for group","Priority","Notes"]
hdr(qc,4,H)
Q=[
("Amazon India","Aurra Hype apparel, Designomics gifting","Free account; pay-per-sale + optional FBA","~10-20% + fees","High — discovery + trust","P1","A+ content + Sponsored Products; use FBA for Prime badge"),
("Flipkart","Apparel + gifting, tier 2/3 reach","Free listing; commission per sale","~5-20%","High","P1","Flipkart Ads + Big Billion Days spikes"),
("Myntra","Aurra Hype fashion-native buyers","Brand approval needed","~25-40% effective","Medium-High","P2","Category-native; higher take rate; strong for streetwear"),
("Ajio","Aurra Hype fashion","Brand onboarding","~high","Medium","P2","Reliance ecosystem; good festive events"),
("Blinkit","Designomics impulse gifting SKUs","~₹25k/SKU/state listing (as ad credits) + ~₹2-3L/mo marketing","15-25% + fees","Medium (needs >65% margin)","P3","Selective onboarding; best for fast-moving low-price gifts"),
("Zepto","Premium gifting positioning","Bundled packages ~₹5-6L to start; Atom analytics ₹30k/mo","10-25%","Low-Medium","P3","Expensive entry; only after product-market fit"),
("Swiggy Instamart","Easiest q-comm entry, South India","No onboarding fee; weekly POs ~₹2-5k","15-25%","Medium","P3","Lightest commitment to test q-commerce"),
("Meesho","Value gifting / stickers volume","Free; 0% commission model","Low","Medium","P2","Great for low-price, high-volume Designomics SKUs"),
("Own D2C (Shopify)","All — highest margin","Owned stack","0% (only gateway ~2%)","Core","P0","Always the priority channel; marketplaces feed it"),
("WhatsApp Commerce","Repeat + gifting + drops","API ~₹5-10k/mo","Per-conversation","High","P1","Catalog + broadcast + checkout; India's highest-converting channel"),
]
r=5; pri_color={"P0":RED,"P1":ORANGE,"P2":TEAL,"P3":"777777"}
for row in Q:
    for ci,v in enumerate(row,1):
        cell=qc.cell(row=r,column=ci,value=v); cell.border=border
        cell.font=Font(name=FONT,size=9,color="222222")
        cell.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if ci==6 else "left"))
    pc=qc.cell(row=r,column=6); pc.font=Font(name=FONT,size=9,bold=True,color="FFFFFF")
    pc.fill=PatternFill("solid",fgColor=pri_color.get(row[5],"999999"))
    if r%2==0:
        for ci in [1,2,3,4,5,7]: qc.cell(row=r,column=ci).fill=PatternFill("solid",fgColor="FAFBFC")
    r+=1
qc.freeze_panes="A5"
for i,w in enumerate([18,26,34,16,22,9,38],1): qc.column_dimensions[get_column_letter(i)].width=w

# order sheets
order=["Overview","Master Stack","AI Automations","90-Day Roadmap","Q-Commerce & Marketplaces","Budget Summary"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.save(P)
print("saved all sheets")
